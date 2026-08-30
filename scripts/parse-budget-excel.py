#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
parse-budget-excel.py — Parse the 2026 Budget Excel into finance/budget.json
and update examples/finance-budget.json mock data for the Shogun web portal.

The Excel is a company-wide P&L budget with:
  - Column A: P&L line item names ( Revenue / Cost of Sales / Expenses sections )
  - Column B: "Jan 1 - May 31 2026 (YTD)" — actual YTD (reference; real actuals come from QBO)
  - Column C: Extrapolated 12 months
  - Columns G–S: Monthly budget figures (Jan–Dec)
  - Column T: TOTAL (annual budget = sum of monthly)

Dashboard row level:
  - Revenue: individual line items (Hardware, Maintenance, etc.)
  - Cost of Sales: individual line items (Cloud, Hardware Costs, etc.)
  - Other Income: individual line items (Grant, MV car instalment)
  - Expenses: category subtotals only (Total for Employee Benefits, Total for Payroll, etc.)
    + standalone non-categorized items (Depreciation, Exchange Gain or Loss, etc.)

Usage:
    python scripts/parse-budget-excel.py <path-to-excel> [--output-budget <path>] [--output-mock <path>] [--ytd-months 5]

Budget source: Excel (monthly columns + T column)
Actual source: QBO (via acct_get_profit_loss) — for dev/mock, column B (YTD actual from Excel) is used.
"""
import argparse
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── Excel column mapping (1-indexed) ──
COL_NAME = 1        # A — line item name
COL_YTD_ACTUAL = 2  # B — "Jan 1 - May 31 2026 (YTD)" — actual YTD reference
COL_TARGET = 5       # E — "Target 2026"
COL_JAN = 7          # G — January
COL_GAP = 14         # N — gap column (no date header, skip)
COL_MONTH_END = 19   # S — December (G=Jan through S=Dec, excluding N gap)
COL_TOTAL = 20       # T — TOTAL (annual budget)

# Section markers — row number → section name
SECTION_MARKERS = {
    5:  "Revenue",
    16: "Cost of Sales",
    33: "Other Income",
    38: "Expenses",
}

# Rows that should be skipped entirely
SKIP_PATTERNS = ["[DO NOT USE]", "Sales of Product Income"]

# Section total rows (grand totals, not category subtotals)
GRAND_TOTAL_NAMES = {
    "TOTAL REVENUE", "TOTAL COST OF SALES", "GROSS PROFIT (GP)",
    "TOTAL OTHER INCOME", "Total for Expenses", "Other Expenses",
    "Net earnings (PBT)", "Taxation (24%)", "Profit after Taxation (PAT)",
}

# Category headers in Expenses — parent rows with no data, just grouping
EXPENSE_CATEGORY_HEADERS = {
    "Advertisement and promotions", "Business Development Costs",
    "Employee Benefits", "Finance costs", "Human Resources",
    "Insurance", "Interest costs", "Office Expenses", "Payroll",
    "Professional Fees", "Rent, Utilities & Phone", "Taxes",
}


def is_indented(name: str) -> bool:
    """Sub-items are indented with 3+ leading spaces (Excel uses 3-space indent)."""
    return name.startswith("   ") or name.startswith("\t")


def is_subtotal(name: str) -> bool:
    """Subtotal rows start with 'Total ' or 'Total for '."""
    stripped = name.strip()
    return stripped.startswith("Total for ") or stripped.startswith("Total ")


def is_section_header(name: str, t_val) -> bool:
    """Section headers are ALL CAPS with no T (TOTAL) value."""
    stripped = name.strip()
    return stripped.isupper() and (t_val is None or t_val == 0) and len(stripped) > 3


def parse_excel(excel_path: str, ytd_months: int = 5) -> list[dict]:
    """Parse the Excel and return a list of P&L line items for the dashboard."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Budget"]

    line_items = []
    current_section = None

    for row in range(1, ws.max_row + 1):
        name = ws.cell(row, COL_NAME).value
        if name is None:
            continue

        name_str = str(name)
        stripped_name = name_str.strip()
        t_val = ws.cell(row, COL_TOTAL).value
        b_val = ws.cell(row, COL_YTD_ACTUAL).value
        e_val = ws.cell(row, COL_TARGET).value

        # Track current section
        if row in SECTION_MARKERS:
            current_section = SECTION_MARKERS[row]
            continue

        # Skip empty names, known skip patterns
        if not stripped_name or any(p in stripped_name for p in SKIP_PATTERNS):
            continue

        # Skip section headers (ALL CAPS, no data)
        if is_section_header(stripped_name, t_val):
            continue

        # Skip grand totals
        if stripped_name in GRAND_TOTAL_NAMES:
            continue

        # ── Section-specific row selection ──
        if current_section == "Revenue":
            # Revenue: all line items are standalone (not indented) — show them all, skip subtotals
            if is_subtotal(stripped_name) or is_indented(name_str):
                continue
        elif current_section == "Cost of Sales":
            # COS: show standalone items (Cloud, Hardware Costs, etc.) + category subtotal rows
            # Skip indented sub-items (Services OTC, Travel, etc. under category headers)
            if is_indented(name_str):
                continue
            # Skip "Total for X" subtotals with no budget (they're just roll-ups of indented items)
            if is_subtotal(stripped_name) and (t_val is None or t_val == 0):
                continue
        elif current_section == "Other Income":
            # Other Income: items are indented but there are no subtotals — include all
            if is_subtotal(stripped_name):
                continue
        elif current_section == "Expenses":
            # Expenses: show only category subtotals + standalone non-indented items
            # Skip indented sub-items and indented category headers (they're 3-space indented)
            if is_indented(name_str):
                continue
            # Skip category headers (parent rows with no data, just grouping labels)
            if stripped_name in EXPENSE_CATEGORY_HEADERS:
                continue
            # Skip "Purchases" and "SST Agency Expense" if they have no budget
            if stripped_name in ("Purchases", "SST Agency Expense") and (t_val is None or t_val == 0):
                continue
        # else: Other section — include everything

        # Get monthly budget values (G through S = Jan through Dec, skip N gap column)
        monthly = []
        for col in range(COL_JAN, COL_MONTH_END + 1):
            if col == COL_GAP:
                continue  # Skip gap column N (no date header)
            v = ws.cell(row, col).value
            monthly.append(float(v) if v is not None else 0.0)

        budget_annual = float(t_val) if t_val is not None else sum(monthly)
        budget_ytd = sum(monthly[:ytd_months])
        actual_ytd_excel = float(b_val) if b_val is not None else 0.0
        target_2026 = float(e_val) if e_val is not None else None

        # Skip rows with no budget and no actual
        if budget_annual == 0 and actual_ytd_excel == 0 and target_2026 is None:
            continue
        # Skip subtotal rows with zero budget (they're roll-ups with no independent budget)
        if is_subtotal(stripped_name) and budget_annual == 0 and target_2026 is None:
            continue

        variance = actual_ytd_excel - budget_ytd
        variance_pct = (variance / budget_ytd * 100.0) if budget_ytd != 0 else 0.0

        line_items.append({
            "section": current_section or "Other",
            "account_name": stripped_name,
            "budget_annual": round(budget_annual, 2),
            "budget_ytd": round(budget_ytd, 2),
            "actual_ytd": round(actual_ytd_excel, 2),
            "variance": round(variance, 2),
            "variance_pct": round(variance_pct, 1),
            "monthly_budget": [round(m, 2) for m in monthly],
            "target_2026": target_2026,
        })

    return line_items


def generate_budget_json(line_items: list[dict], ytd_months: int = 5) -> dict:
    """Generate the canonical finance/budget.json for bva-variance-analysis skill."""
    section_prefix = {"Revenue": "4", "Cost of Sales": "5", "Other Income": "6", "Expenses": "7"}
    coded_lines = []
    for i, item in enumerate(line_items):
        prefix = section_prefix.get(item["section"], "8")
        coded_lines.append({
            "account_code": f"{prefix}{i+100:03d}",
            "account_name": item["account_name"],
            "section": item["section"],
            "budget_amount": item["budget_annual"],
            "budget_ytd": item["budget_ytd"],
            "monthly_budget": item.get("monthly_budget", []),
            "driver": "",
            "notes": "From 2026 Budget Excel",
        })

    return {
        "period": "2026",
        "year": 2026,
        "currency": "MYR",
        "company": "",
        "ytd_months": ytd_months,
        "ytd_period": f"Jan 1 – May 31 2026 (YTD, {ytd_months} months)" if ytd_months == 5 else f"YTD ({ytd_months} months)",
        "budget_source": "2026 Budget (v3).xlsx",
        "actuals_source": "QuickBooks Online (QBO) — via acct_get_profit_loss MCP tool",
        "lines": coded_lines,
    }


def generate_dashboard_mock(line_items: list[dict]) -> list[dict]:
    """Generate the bvaLineItems array for examples/finance-budget.json dashboard_mock."""
    section_order = ["Revenue", "Cost of Sales", "Other Income", "Expenses"]
    grouped = {}
    for item in line_items:
        s = item["section"]
        if s not in grouped:
            grouped[s] = []
        grouped[s].append({
            "section": s,
            "account_name": item["account_name"],
            "budget_annual": item["budget_annual"],
            "budget_ytd": item["budget_ytd"],
            "actual_ytd": item["actual_ytd"],
            "variance": item["variance"],
            "variance_pct": item["variance_pct"],
            "monthly_budget": item.get("monthly_budget", []),
        })

    result = []
    for s in section_order:
        if s in grouped:
            result.extend(grouped[s])
    return result


def update_examples_mock(mock_path: str, bva_line_items: list[dict]) -> None:
    """Update examples/finance-budget.json budget lines, then rebuild the whole
    dashboard_mock via scripts/generate_finance_mock.py so every tab stays
    internally aligned (BvA, trends, cash, AR/AP, balance sheet)."""
    if not os.path.exists(mock_path):
        print(f"WARNING: {mock_path} does not exist — skipping mock update")
        return

    with open(mock_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Update the canonical lines + budget_baseline
    budget_data = generate_budget_json(bva_line_items)
    data["lines"] = budget_data["lines"]
    data["budget_baseline"] = {}
    for line in budget_data["lines"]:
        data["budget_baseline"][line["account_name"]] = line["budget_amount"]

    with open(mock_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print(f"Updated {mock_path} budget lines ({len(bva_line_items)} P&L items)")

    # Regenerate the aligned mock ledger from the new budget
    try:
        from generate_finance_mock import build, validate, TARGET
        rebuilt = build(json.loads(json.dumps(data)))
        errs = validate(rebuilt)
        if errs:
            print(f"WARNING: aligned rebuild failed invariants: {errs}")
            return
        data["dashboard_mock"] = rebuilt
        with open(mock_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
            f.write("\n")
        print("Rebuilt aligned dashboard_mock (all tabs reconciled)")
    except Exception as e:
        print(f"WARNING: aligned rebuild skipped: {e}")
        print("  → run `python scripts/generate_finance_mock.py` to fix")


def main():
    parser = argparse.ArgumentParser(description="Parse 2026 Budget Excel into budget.json + dashboard mock data")
    parser.add_argument("excel_path", help="Path to the 2026 Budget Excel file")
    parser.add_argument("--output-budget", default="finance/budget.json",
                        help="Output path for finance/budget.json (default: finance/budget.json)")
    parser.add_argument("--output-mock", default="examples/finance-budget.json",
                        help="Output path for examples/finance-budget.json (default: examples/finance-budget.json)")
    parser.add_argument("--ytd-months", type=int, default=5,
                        help="Number of completed months for YTD budget calculation (default: 5 = Jan-May)")
    args = parser.parse_args()

    if not os.path.exists(args.excel_path):
        print(f"ERROR: Excel file not found: {args.excel_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Parsing Excel: {args.excel_path}")
    print(f"YTD months: {args.ytd_months} (Jan through {['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][args.ytd_months-1]})")

    line_items = parse_excel(args.excel_path, args.ytd_months)

    # Print summary
    sections = {}
    for item in line_items:
        s = item["section"]
        if s not in sections:
            sections[s] = {"count": 0, "budget_annual": 0, "budget_ytd": 0, "actual_ytd": 0}
        sections[s]["count"] += 1
        sections[s]["budget_annual"] += item["budget_annual"]
        sections[s]["budget_ytd"] += item["budget_ytd"]
        sections[s]["actual_ytd"] += item["actual_ytd"]

    print(f"\n{'Section':<20} {'Lines':>6} {'Budget Annual':>15} {'Budget YTD':>15} {'Actual YTD':>15}")
    print("-" * 75)
    for s, v in sections.items():
        print(f"{s:<20} {v['count']:>6} {v['budget_annual']:>15,.2f} {v['budget_ytd']:>15,.2f} {v['actual_ytd']:>15,.2f}")
    print("-" * 75)
    total_budget = sum(v["budget_annual"] for v in sections.values())
    total_ytd = sum(v["budget_ytd"] for v in sections.values())
    total_actual = sum(v["actual_ytd"] for v in sections.values())
    print(f"{'TOTAL':<20} {len(line_items):>6} {total_budget:>15,.2f} {total_ytd:>15,.2f} {total_actual:>15,.2f}")

    # Generate budget.json
    budget_data = generate_budget_json(line_items, args.ytd_months)
    budget_dir = os.path.dirname(args.output_budget)
    if budget_dir:
        os.makedirs(budget_dir, exist_ok=True)
    with open(args.output_budget, "w", encoding="utf-8") as f:
        json.dump(budget_data, f, indent=2, ensure_ascii=False)
    print(f"\nWrote budget.json → {args.output_budget} ({len(line_items)} lines)")

    # Update examples/finance-budget.json mock data
    dashboard_items = generate_dashboard_mock(line_items)
    update_examples_mock(args.output_mock, dashboard_items)

    print(f"\n✅ Parsed {len(line_items)} P&L line items across {len(sections)} sections")
    print(f"   Budget source: Excel (2026 Budget v3)")
    print(f"   Actuals: fictional demo ledger (generated by generate_finance_mock.py)")


if __name__ == "__main__":
    main()
