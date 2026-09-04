#!/usr/bin/env python3
"""cashflow-to-snapshot.py

Fill the Cash Flow tab's empty charts with REAL data + a labeled plan:

1. monthly_pl_trend  — actual P&L per month from each management report's
                       P&L sheet (June + any later reports found). Revenue,
                       expenses (total), net earnings — real numbers.
2. burn_trend        — monthly total expenses from same P&L sheets.
3. cash_flow_trend   — monthly net cash from operating activities + end cash
                       from each report's CF sheet.
4. cash_flow_forecast — 12-month projection from the 2026 budget plan
                       (revenue − COS − expenses + other income), fanned
                       low/high ±20%. Labeled source: 2026-budget-v3.xlsx.
5. dso               — AR / (YTD revenue / days) using the AR snapshot.

Writes finance/snapshots/cash + finance/snapshots/pl. Idempotent.
"""
import base64, json, os, re, subprocess
from datetime import datetime
import openpyxl

PWFILE = "/home/tapway/brain/finance/202606-management-report.xlsx"
BUDGET_XLSX = "/home/tapway/brain/finance/2026-budget-v3.xlsx"
PGPW = os.environ.get("GBRAIN_PG_PASSWORD", base64.b64decode("aGVybWVzX3Mzc3Npb25zXzIwMjY=").decode())
MONTHS = ["2026-01","2026-02","2026-03","2026-04","2026-05","2026-06",
          "2026-07","2026-08","2026-09","2026-10","2026-11","2026-12"]
MONTH_LBL = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]


def pg(sql):
    env = dict(os.environ); env["PGPASSWORD"] = PGPW
    p = subprocess.run(["psql","-h","127.0.0.1","-U","hermes","-d","gbrain","-c",sql],
                       capture_output=True, text=True, env=env, timeout=60)
    if p.returncode != 0:
        raise SystemExit(p.stderr[-400:])


def get_snap(slug):
    env = dict(os.environ); env["PGPASSWORD"] = PGPW
    p = subprocess.run(["psql","-h","127.0.0.1","-U","hermes","-d","gbrain","-t","-A","-c",
                        f"SELECT frontmatter::text FROM pages WHERE source_id='default' AND slug='{slug}'"],
                       capture_output=True, text=True, env=env, timeout=30)
    try:
        return json.loads(p.stdout.strip())
    except Exception:
        return {}


def put_snap(slug, title, data):
    fm = json.dumps(data, default=str).replace("'", "''")
    body = json.dumps(data, indent=2, default=str).replace("'", "''")
    pg(f"""
    INSERT INTO pages (source_id, slug, type, page_kind, title, frontmatter, compiled_truth,
                       content_hash, created_at, updated_at)
    VALUES ('default', '{slug}', 'finance', 'markdown', '{title}',
            '{fm}'::jsonb, '{body}', '', now(), now())
    ON CONFLICT (source_id, slug) DO UPDATE SET
        frontmatter = EXCLUDED.frontmatter,
        compiled_truth = EXCLUDED.compiled_truth,
        title = EXCLUDED.title,
        updated_at = now();
    """)


def find_reports():
    found = []
    base = os.path.dirname(PWFILE)
    for root, _d, files in os.walk(base):
        for f in sorted(files):
            if "management-report" in f.lower() and f.lower().endswith(".xlsx"):
                found.append(os.path.join(root, f))
    return found


def month_tag(path):
    m = re.search(r"(\d{4})(\d{2})", os.path.basename(path))
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m")


def pl_totals(path):
    """(revenue, total_expenses, net) from a report's P&L sheet."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "P&L" not in wb.sheetnames:
        return None
    ws = wb["P&L"]
    rev = exp = net = None
    for r in ws.iter_rows(values_only=True):
        name = str(r[0] or "").strip().lower()
        if name in ("total revenue", "total income", "total for income"):
            if rev is None: rev = r[1]
        elif name == "total for expenses":
            if exp is None: exp = r[1]
        elif name == "net earnings":
            if net is None: net = r[1]
    return (rev, exp, net)


def cf_totals(path):
    """(net_op_cash, end_cash) from a report's CF sheet."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "CF" not in wb.sheetnames:
        return None
    ws = wb["CF"]
    op = end = None
    for r in ws.iter_rows(values_only=True):
        name = str(r[0] or "").strip().lower()
        if "net cash from operating" in name:
            op = r[1]
        elif "end of year" in name:
            end = r[1]
    return (op, end)


def budget_monthly_net():
    """Monthly planned net cash (revenue − COS − expenses + other income)."""
    wb = openpyxl.load_workbook(BUDGET_XLSX, read_only=True, data_only=True)
    rows = [list(r) for r in wb["Budget"].iter_rows(values_only=True)]
    lookup = {}
    for r in rows:
        lookup[str(r[0] or "").strip().lower()] = r
    def ser(name):
        r = lookup.get(name)
        return [float(v) if isinstance(v, (int, float)) else 0.0 for v in r[6:19]] if r else [0.0]*12
    rev = ser("total revenue")
    cos = ser("total cost of sales")
    oth = ser("total other income")
    exp = ser("total for expenses")
    return [round(rev[i] - cos[i] + oth[i] - exp[i], 2) for i in range(12)]


def main():
    # ── 1 & 2 & 3: gather real monthly points from every report ──
    trend = []   # monthly_pl_trend
    burn = []    # burn_trend
    cf_flow = [] # cash_flow_trend
    for path in find_reports():
        tag = month_tag(path)
        p = pl_totals(path)
        if p and p[0] is not None:
            rev, exp, net = (float(x) if x is not None else 0.0 for x in p)
            trend.append({"month": tag, "revenue": round(rev, 2),
                          "expenses": round(exp, 2), "net_profit": round(net, 2)})
            burn.append({"month": tag, "burn": round(exp, 2)})
        c = cf_totals(path)
        if c and c[0] is not None:
            cf_flow.append({"month": tag, "net_op": round(float(c[0]), 2),
                            "end_cash": round(float(c[1] or 0), 2),
                            "cash": round(float(c[1] or 0), 2),
                            "netFlow": round(float(c[0]), 2)})
    trend.sort(key=lambda x: x["month"])
    burn.sort(key=lambda x: x["month"])
    cf_flow.sort(key=lambda x: x["month"])

    # ── 4: budget-plan forecast (12 months, fanned) ──
    net_by_month = budget_monthly_net()
    # seed: latest actual end_cash (or bank balance from cash snapshot)
    cash_snap = get_snap("finance/snapshots/cash")
    latest_cash = cash_snap.get("total_liquid_cash") or 0.0
    if cf_flow and cf_flow[-1].get("end_cash"):
        latest_cash = cf_flow[-1]["end_cash"]
    # find index of the latest actual month in the budget year
    start_i = 0
    if trend:
        last = trend[-1]["month"]  # e.g. 2026-07
        try:
            start_i = MONTHS.index(last) + 1
        except ValueError:
            start_i = 6  # default: forecast from Jul
    forecast = []
    running = latest_cash
    for i in range(start_i, 12):
        running += net_by_month[i]
        total = round(running, 2)
        forecast.append({
            "month": MONTH_LBL[i],
            "total": total,
            "low": round(total * 0.8, 2),
            "high": round(total * 1.2, 2),
        })

    dso_src = get_snap("finance/snapshots/ar")
    ar_total = float(dso_src.get("total_ar") or 0)
    pl_snap = get_snap("finance/snapshots/pl")
    rev_ytd = float(pl_snap.get("revenue_ytd") or 0)
    dso = round(ar_total / (rev_ytd / 181), 1) if rev_ytd else 0.0

    # ── write back into cash + pl snapshots ──
    cash_snap["burn_trend"] = burn
    cash_snap["cash_flow_trend"] = cf_flow
    cash_snap["cash_flow_forecast"] = forecast
    cash_snap["cash_flow_breakdown"] = {
        "operating": sum(x["net_op"] for x in cf_flow),
        "investing": 0.0,
        "financing": 0.0,
        "net_change": round(sum(x["net_op"] for x in cf_flow), 2),
    }
    cash_snap["forecast_source"] = "2026-budget-v3.xlsx (plan)"
    put_snap("finance/snapshots/cash", "Finance Snapshot — Cash (2026)", cash_snap)

    pl_snap["monthly_pl_trend"] = trend
    put_snap("finance/snapshots/pl", "Finance Snapshot — P&L (2026)", pl_snap)
    if dso_src:
        dso_src["dso"] = dso
        put_snap("finance/snapshots/ar", "Finance Snapshot — AR (2026)", dso_src)

    print(f"✓ cash + pl snapshots updated")
    print(f"  monthly_pl_trend: {len(trend)} real points — " + ", ".join(f"{p['month']}: rev {p['revenue']:,.0f} / exp {p['expenses']:,.0f} / net {p['net_profit']:,.0f}" for p in trend))
    print(f"  burn_trend:       {len(burn)} points — " + ", ".join(f"{b['month']}: {b['burn']:,.0f}" for b in burn))
    print(f"  cash_flow_trend:  {len(cf_flow)} points — " + ", ".join(f"{c['month']}: op {c['net_op']:,.0f} → end {c['end_cash']:,.0f}" for c in cf_flow))
    print(f"  cash_flow_forecast: {len(forecast)} months (plan-based, fan ±20%) — " + ", ".join(f"{f['month']}: {f['total']:,.0f}" for f in forecast[:4]) + " …")
    print(f"  dso:              {dso} days (AR {ar_total:,.0f} / revYTD {rev_ytd:,.0f})")


if __name__ == "__main__":
    main()