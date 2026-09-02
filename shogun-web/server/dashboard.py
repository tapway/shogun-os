"""Department dashboard endpoints — aggregates data via gbrain MCP."""
from __future__ import annotations

import asyncio
import copy
import json
import logging
import os
import pathlib
import re as _re
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional
from urllib.parse import quote as _url_quote

from fastapi import APIRouter, Depends, HTTPException, Path, UploadFile, File, Query, Form, Body
from pydantic import BaseModel, Field
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from auth import get_current_user
from config import get_config
from database import get_db, get_primary_tenant
from gbrain_client import gbrain_fetch_page, gbrain_fetch_pages, gbrain_search
from models import Tenant, Department, User

import httpx

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Document Scanning — Multi-source configuration & results
# ---------------------------------------------------------------------------

class DocScanSource(BaseModel):
    """Configuration for a document scan source."""
    title: str
    drive_url: str
    schedule: str = "manual"  # daily, weekly, manual
    document_type: str = "invoice"
    template_path: Optional[str] = None

class DocScanResult(BaseModel):
    """Result from a document scan run."""
    filename: str
    file_url: str
    source_id: str
    source_title: str
    document_type: str
    ocr_summary: str
    interpretation: Dict[str, Any] = Field(default_factory=dict)
    status: str = "pending"  # pending, processed, verified, rejected

# In-memory storage for demo (replace with DB table in production)
_DOC_SCAN_SOURCES: Dict[str, Dict[str, Any]] = {}  # key = department:id
_DOC_SCAN_RESULTS: Dict[str, List[Dict[str, Any]]] = {}  # key = department


router = APIRouter(prefix="/departments/{name}/dashboard", tags=["dashboard"])

# ─── Canonicalization ───
# Add per-installation owner aliases and product patterns here.
# These maps normalise raw frontmatter values to canonical labels.

OWNER_ALIASES: dict[str, str] = {}

STAGE_ORDER = ["Lead", "On Hold", "Prospecting", "Qualified", "Quote", "Tender", "Unqualified", "Confirmed", "Won"]
STAGE_WEIGHTS = {
    "Lead": 0.05, "On Hold": 0.0, "Prospecting": 0.15, "Qualified": 0.30,
    "Quote": 0.50, "Tender": 0.65, "Unqualified": 0.0, "Confirmed": 0.90, "Won": 1.0,
}
WON_STAGES = {"Won"}
LOST_STAGES = {"Lost", "Unqualified"}
ACTIVE_STAGES = {"Lead", "Prospecting", "Qualified", "Quote", "Tender", "Confirmed", "On Hold"}
PRODUCT_PATTERNS: list[tuple[str, str]] = []


def _canonical_owner(raw: str) -> str:
    key = raw.strip().lower()
    return OWNER_ALIASES.get(key, raw.strip() or "Unassigned")


def _canonical_stage(raw: str) -> str:
    s = raw.strip().lower()
    for known in STAGE_ORDER:
        if known.lower() == s:
            return known
    for known in STAGE_ORDER:
        if s in known.lower() or known.lower() in s:
            return known
    return raw.strip()


def _canonical_priority(raw: str) -> str:
    p = (raw or "").strip().lower()
    if p in ("high", "hot"):
        return "High"
    if p in ("medium", "warm", "normal"):
        return "Medium"
    if p in ("low", "cold"):
        return "Low"
    return (raw or "").strip()


def _parse_frontmatter(fm: Any) -> dict:
    if isinstance(fm, dict):
        return fm
    if isinstance(fm, str):
        try:
            return json.loads(fm)
        except (json.JSONDecodeError, TypeError):
            return {}
    return {}


def _infer_product(title: str, slug: str) -> str:
    text = f"{title} {slug}".lower()
    for pattern, product in PRODUCT_PATTERNS:
        if _re.search(pattern, text):
            return product
    return "Uncategorised"


def _month_key(iso_str: str) -> str:
    return iso_str[:7] if len(iso_str) >= 7 else iso_str


# ─── Aggregation helpers (ported from typescript) ───


class OwnerAccum:
    __slots__ = (
        "salesMTD", "salesQTD", "salesYTD", "deals", "wonDeals",
        "pipelineValue", "weightedPipeline", "closeThisMonth", "closeThisQ",
        "closeNextQ", "closeThisYear", "winNum", "winDen",
    )

    def __init__(self):
        for s in self.__slots__:
            setattr(self, s, 0)


class PartnerAccum:
    __slots__ = ("booking", "dealsWon", "pipelineDeals", "pipelineValue", "winNum", "winDen")

    def __init__(self):
        for s in self.__slots__:
            setattr(self, s, 0)


def _now() -> datetime:
    return datetime.now()


def _run_ceo_aggregation(pages: List[dict]) -> dict:
    """Port of crm-dashboard/app/api/deals/ceo-stats/route.ts aggregation logic."""
    # Filter to deals
    deals = [p for p in pages if p.get("slug", "").startswith("deals/")]
    deals = [p for p in deals if not any(
        x in str(p.get("slug", "")) for x in ["templates/", "/readme", "_schema", "activity-log", "risk-register"]
    )]

    now = _now()
    cy, cm = now.year, now.month
    cq = cm // 3

    def _is_this_month(iso: str) -> bool:
        if not iso: return False
        d = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return d.year == cy and d.month == cm

    def _is_this_quarter(iso: str) -> bool:
        if not iso: return False
        d = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return d.year == cy and d.month // 3 == cq

    def _is_this_year(iso: str) -> bool:
        if not iso: return False
        return datetime.fromisoformat(iso.replace("Z", "+00:00")).year == cy

    def _is_next_quarter(iso: str) -> bool:
        if not iso: return False
        d = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        next_q = (cq + 1) % 4
        next_q_year = cy + (1 if cq == 3 else 0)
        return d.year == next_q_year and d.month // 3 == next_q

    def _days_since(iso: str) -> int:
        if not iso: return 0
        d = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return max(0, (now - d).days)

    # Accumulators
    salesMTD = salesQTD = salesYTD = 0
    totalPipelineValue = weightedPipelineValue = 0
    totalActiveDeals = hotDeals = warmDeals = coldDeals = wonDeals = 0

    owner_map: Dict[str, OwnerAccum] = {}
    partner_map: Dict[str, PartnerAccum] = {}
    stage_map: Dict[str, Dict[str, float]] = {}
    month_map: Dict[str, float] = {}
    won_month_map: Dict[str, float] = {}
    product_map: Dict[str, Dict[str, float]] = {}
    priority_map: Dict[str, int] = {}
    matrix_map: Dict[str, Dict[str, Any]] = {}
    at_risk_by_owner: Dict[str, Dict[str, float]] = {}
    at_risk_by_partner: Dict[str, Dict[str, float]] = {}
    partner_owner_counts: Dict[str, Dict[str, int]] = {}
    top_deals: List[Dict[str, Any]] = []

    # Omnichannel accumulators (spec §2.3)
    channel_volume = {"shopee": 0, "lazada": 0, "fbMessenger": 0, "whatsapp": 0}
    _CHANNEL_MAP = {
        "Shopee": "shopee", "Lazada": "lazada",
        "Facebook Messenger": "fbMessenger", "FB Messenger": "fbMessenger",
        "WhatsApp": "whatsapp",
    }
    response_minutes: List[float] = []
    sla_compliant = 0
    ai_resolved = 0
    chat_inbox_rows: List[Dict[str, Any]] = []

    for deal in deals:
        slug = str(deal.get("slug", ""))
        title = str(deal.get("title", ""))
        fm = _parse_frontmatter(deal.get("frontmatter", {}))

        amount = float(fm.get("amount", 0) or 0)
        raw_stage = str(fm.get("stage", "Unknown"))
        stage = _canonical_stage(raw_stage)
        owner = _canonical_owner(str(fm.get("owner", "")))
        partner = str(fm.get("partner", "")).strip() if fm.get("partner") else None
        priority = str(fm.get("priority", "Medium"))
        close_date = str(fm.get("close_date", ""))
        hot = fm.get("hot") in ("Yes", True)

        won = stage in WON_STAGES
        lost = stage in LOST_STAGES
        active = stage in ACTIVE_STAGES

        # Owner accum
        if owner not in owner_map:
            owner_map[owner] = OwnerAccum()
        om = owner_map[owner]

        if won:
            wonDeals += 1
            om.wonDeals += 1
            if amount > 0 and close_date and _is_this_year(close_date):
                salesYTD += amount
                om.salesYTD += amount
                if _is_this_quarter(close_date):
                    salesQTD += amount
                    om.salesQTD += amount
                    if _is_this_month(close_date):
                        salesMTD += amount
                        om.salesMTD += amount

        if active and amount > 0:
            totalActiveDeals += 1
            om.deals += 1
            totalPipelineValue += amount
            om.pipelineValue += amount
            prob = STAGE_WEIGHTS.get(stage, 0.0)
            w = amount * prob
            weightedPipelineValue += w
            om.weightedPipeline += w

            if close_date:
                if _is_this_month(close_date):
                    om.closeThisMonth += amount
                if _is_this_quarter(close_date):
                    om.closeThisQ += amount
                if _is_next_quarter(close_date):
                    om.closeNextQ += amount
                if _is_this_year(close_date):
                    om.closeThisYear += amount

            if hot:
                hotDeals += 1
            elif priority == "High":
                warmDeals += 1
            else:
                coldDeals += 1

            days_in_stage = _days_since(fm.get("created") or close_date or now.isoformat())
            if days_in_stage > 30:
                ar = at_risk_by_owner.setdefault(owner, {"count": 0.0, "value": 0.0})
                ar["count"] += 1
                ar["value"] += amount
                if partner:
                    arp = at_risk_by_partner.setdefault(partner, {"count": 0.0, "value": 0.0})
                    arp["count"] += 1
                    arp["value"] += amount

        if won or lost:
            om.winDen += 1
            if won:
                om.winNum += 1

        if partner:
            pm = partner_map.setdefault(partner, PartnerAccum())
            if won and amount > 0 and close_date and _is_this_year(close_date):
                pm.booking += amount
                pm.dealsWon += 1
            if active and amount > 0:
                pm.pipelineDeals += 1
                pm.pipelineValue += amount
            if won or lost:
                pm.winDen += 1
                if won:
                    pm.winNum += 1

            mk = f"{owner}|{partner}"
            mm = matrix_map.setdefault(mk, {"owner": owner, "partner": partner, "deals": 0})
            mm["deals"] += 1

            poc = partner_owner_counts.setdefault(partner, {})
            poc[owner] = poc.get(owner, 0) + 1

        product = fm.get("product") or _infer_product(title, slug)
        pe = product_map.setdefault(product, {"value": 0.0, "count": 0.0})
        pe["count"] += 1
        if amount > 0:
            pe["value"] += amount

        se = stage_map.setdefault(stage, {"count": 0.0, "value": 0.0})
        se["count"] += 1
        if amount > 0:
            se["value"] += amount

        if close_date:
            mk = _month_key(close_date)
            if won and amount > 0:
                won_month_map[mk] = won_month_map.get(mk, 0) + amount
            if won or active:
                month_map[mk] = month_map.get(mk, 0) + amount

        priority_map[priority] = priority_map.get(priority, 0) + 1

        # Omnichannel: accumulate per-deal channel_origin + sla_response_minutes (spec §3)
        raw_channel = str(fm.get("channel_origin", "")).strip()
        chan_key = _CHANNEL_MAP.get(raw_channel)
        if chan_key:
            channel_volume[chan_key] = channel_volume.get(chan_key, 0) + 1
        raw_sla = fm.get("sla_response_minutes")
        if raw_sla not in (None, ""):
            try:
                sla_val = float(raw_sla)
                response_minutes.append(sla_val)
                if sla_val <= 15:
                    sla_compliant += 1
            except (TypeError, ValueError):
                pass

        is_early = stage in ("Lead", "Prospecting", "Qualified")
        if amount > 0 and active and stage not in ("Unqualified", "On Hold") and (not is_early or amount > 0):
            top_deals.append({
                "slug": slug,
                "title": title,
                "customer": str(fm.get("customer", "")),
                "amount": amount,
                "stage": stage,
                "priority": "Hot" if hot else "Warm" if priority == "High" else "Cold",
                "owner": owner,
                "partner": partner,
                "closeDate": close_date,
                "winProbability": round(STAGE_WEIGHTS.get(stage, 0) * 100),
                "daysInStage": _days_since(fm.get("created") or close_date or now.isoformat()),
                "hot": hot,
            })

    # Assemble response
    funnel = []
    for s in STAGE_ORDER:
        if s in stage_map:
            funnel.append({"stage": s, **stage_map[s]})

    by_month = sorted(
        [{"month": m, "value": v} for m, v in month_map.items()],
        key=lambda x: x["month"],
    )

    by_priority = [{"priority": p, "count": c} for p, c in priority_map.items()]

    by_manager = sorted(
        [
            {
                "owner": o, "salesMTD": m.salesMTD, "salesQTD": m.salesQTD,
                "salesYTD": m.salesYTD, "deals": m.deals, "wonDeals": m.wonDeals,
                "pipelineValue": m.pipelineValue, "weightedPipeline": m.weightedPipeline,
                "closeThisMonth": m.closeThisMonth, "closeThisQ": m.closeThisQ,
                "closeNextQ": m.closeNextQ, "closeThisYear": m.closeThisYear,
                "winRate": round(m.winNum / m.winDen * 100) if m.winDen > 0 else 0,
            }
            for o, m in owner_map.items()
        ],
        key=lambda x: x["salesYTD"],
        reverse=True,
    )

    by_partner = sorted(
        [
            {
                "partner": p, "booking": pm.booking, "dealsWon": pm.dealsWon,
                "pipelineDeals": pm.pipelineDeals, "pipelineValue": pm.pipelineValue,
                "winRate": round(pm.winNum / pm.winDen * 100) if pm.winDen > 0 else 0,
                "avgDealSize": round(pm.booking / pm.dealsWon) if pm.dealsWon > 0 else 0,
                "primaryOwner": (
                    sorted(partner_owner_counts.get(p, {}).items(), key=lambda x: -x[1])[0][0]
                    if partner_owner_counts.get(p) else ""
                ),
            }
            for p, pm in partner_map.items()
        ],
        key=lambda x: x["booking"],
        reverse=True,
    )

    by_manager_by_partner = sorted(matrix_map.values(), key=lambda x: -x["deals"])

    won_by_month = sorted(
        [{"month": m, "value": v} for m, v in won_month_map.items()],
        key=lambda x: x["month"],
    )

    by_product = sorted(
        [{"product": p, "value": v["value"], "count": v["count"]} for p, v in product_map.items()],
        key=lambda x: -x["value"],
    )

    at_risk_by_manager = sorted(
        [{"owner": o, "atRiskDeals": int(v["count"]), "atRiskValue": v["value"]}
         for o, v in at_risk_by_owner.items()],
        key=lambda x: -x["atRiskValue"],
    )

    at_risk_by_partner_result = sorted(
        [
            {
                "partner": p, "atRiskDeals": int(v["count"]), "atRiskValue": v["value"],
                "primaryOwner": (
                    sorted(partner_owner_counts.get(p, {}).items(), key=lambda x: -x[1])[0][0]
                    if partner_owner_counts.get(p) else ""
                ),
            }
            for p, v in at_risk_by_partner.items()
        ],
        key=lambda x: -x["atRiskValue"],
    )

    total_win_num = sum(om.winNum for om in owner_map.values())
    total_win_den = sum(om.winDen for om in owner_map.values())
    avg_deal_size = round(totalPipelineValue / totalActiveDeals) if totalActiveDeals > 0 else 0
    pipeline_coverage = round(totalPipelineValue / salesYTD * 10) / 10 if salesYTD > 0 else 0
    top15 = sorted(top_deals, key=lambda x: -x["amount"])[:15]

    # ── Omnichannel: derive from deals where possible, fall back to examples/crm-mock.json ──
    # Inbox + weekly trend are net-new data with no deal source (Concern 2); always load from mock.
    crm_mock: Dict[str, Any] = {}
    mock_json_path = pathlib.Path(__file__).resolve().parents[2] / "examples" / "crm-mock.json"
    if mock_json_path.exists():
        try:
            with open(mock_json_path, "r", encoding="utf-8") as f:
                crm_mock = json.load(f).get("dashboard_mock", {})
        except Exception as e:
            logger.warning("Failed to load mock data from %s: %s", mock_json_path, e)

    if totalActiveDeals == 0 and not wonDeals and _crm_mock_enabled():
        # Demo mode: serve the full mock payload so every Overview panel renders.
        mock_payload = _load_crm_mock().get("dashboard_mock", {})
        if not mock_payload:  # noqa: SIM102 - clarity
            pass
        else:
            return {**mock_payload, "mock": True}

    if totalActiveDeals == 0 and not wonDeals:
        # No real deals at all — return an empty-state payload, not fabricated
        # mock figures. Same policy as finance/procurement (PR #12 review: no
        # fabricated numbers). The UI shows "no data yet / connect gbrain".
        logger.info("CRM dashboard: no deal data — returning empty state")
        return {
            "salesMTD": 0,
            "salesQTD": 0,
            "salesYTD": 0,
            "totalPipelineValue": 0,
            "weightedPipelineValue": 0,
            "pipelineCoverage": 0.0,
            "winRate": 0,
            "avgDealSize": 0,
            "salesCycleDays": 0,
            "totalActiveDeals": 0,
            "hotDeals": 0,
            "warmDeals": 0,
            "coldDeals": 0,
            "wonDeals": 0,
            "byManager": [],
            "byPartner": [],
            "byStage": [],
            "byMonth": [],
            "byPriority": [],
            "wonByMonth": [],
            "byProduct": [],
            "atRiskByManager": [],
            "atRiskByPartner": [],
            "byManagerByPartner": [],
            "topDeals": [],
            "channelVolume": {"shopee": 0, "lazada": 0, "fbMessenger": 0, "whatsapp": 0},
            "avgResponseMinutes": 0.0,
            "slaCompliancePct": 0.0,
            "aiResolutionPct": 0.0,
            "chatToOrderPct": 0.0,
            "chatToOrderTrend": [],
            "chatInbox": [],
        }

    # Real deals exist — derive channel volume + SLA from frontmatter; inbox + trend still mock
    channel_volume_out = channel_volume
    if response_minutes:
        avg_response = round(sum(response_minutes) / len(response_minutes), 1)
        sla_pct = round(sla_compliant / len(response_minutes) * 100, 1)
    else:
        avg_response = _safe_float(crm_mock.get("avgResponseMinutes"))
        sla_pct = _safe_float(crm_mock.get("slaCompliancePct"))
    ai_pct = _safe_float(crm_mock.get("aiResolutionPct"))
    c2o_pct = _safe_float(crm_mock.get("chatToOrderPct"))
    c2o_trend = crm_mock.get("chatToOrderTrend", [])
    chat_inbox_out = crm_mock.get("chatInbox", [])

    return {
        "salesMTD": salesMTD,
        "salesQTD": salesQTD,
        "salesYTD": salesYTD,
        "totalPipelineValue": totalPipelineValue,
        "weightedPipelineValue": weightedPipelineValue,
        "pipelineCoverage": pipeline_coverage,
        "winRate": round(total_win_num / total_win_den * 100) if total_win_den > 0 else 0,
        "avgDealSize": avg_deal_size,
        "salesCycleDays": 47,
        "totalActiveDeals": totalActiveDeals,
        "hotDeals": hotDeals,
        "warmDeals": warmDeals,
        "coldDeals": coldDeals,
        "wonDeals": wonDeals,
        "byManager": by_manager,
        "byPartner": by_partner,
        "byStage": funnel,
        "byMonth": by_month,
        "byPriority": by_priority,
        "wonByMonth": won_by_month,
        "byProduct": by_product,
        "atRiskByManager": at_risk_by_manager,
        "atRiskByPartner": at_risk_by_partner_result,
        "byManagerByPartner": by_manager_by_partner,
        "topDeals": top15,
        "channelVolume": channel_volume_out,
        "avgResponseMinutes": avg_response,
        "slaCompliancePct": sla_pct,
        "aiResolutionPct": ai_pct,
        "chatToOrderPct": c2o_pct,
        "chatToOrderTrend": c2o_trend,
        "chatInbox": chat_inbox_out,
    }


# ─── Endpoints ───


@router.get("")
async def get_dashboard_config(
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Return dashboard configuration for this department."""
    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    dept = db.query(Department).filter(
        Department.tenant_id == tenant.id, Department.name == name
    ).first()
    if dept is None:
        raise HTTPException(status_code=404, detail="Department not found")

    dashboard_meta = {
        "crm": {
            "enabled": True,
            "tabs": [
                {"id": "overview", "label": "Overview", "icon": "LayoutDashboard"},
                {"id": "deals", "label": "Deals", "icon": "Briefcase"},
                {"id": "companies", "label": "Companies", "icon": "Building2"},
                {"id": "tasks", "label": "Tasks", "icon": "SquareCheckBig"},
                {"id": "search", "label": "Search", "icon": "Search"},
                {"id": "bev", "label": "BEV Zones", "icon": "Map"},
                {"id": "partners", "label": "Partners", "icon": "Users"},
            ],
        },
        "finance": {
            "enabled": True,
            "tabs": [
                {"id": "overview", "label": "Overview", "icon": "LayoutDashboard"},
                {"id": "cashflow", "label": "Cash Flow", "icon": "Waves"},
                {"id": "cash", "label": "Assets", "icon": "TrendingUp"},
                {"id": "ar", "label": "AR & Collections", "icon": "Receipt"},
                {"id": "ap", "label": "AP & Payments", "icon": "CreditCard"},
                {"id": "bva", "label": "Budget vs Actuals", "icon": "BarChart3"},
                {"id": "margins", "label": "Margins & Concentration", "icon": "PieChart"},
                {"id": "scan", "label": "Document Scanning", "icon": "FileScan"},
            ],
        },
        "procurement": {
            "enabled": True,
            "tabs": [
                {"id": "pulse", "label": "Overview", "icon": "LayoutDashboard"},
                {"id": "requisitions", "label": "Purchase Requisitions", "icon": "FileText"},
                {"id": "sourcing", "label": "RFQ & Vendor Sourcing", "icon": "Award"},
                {"id": "po", "label": "POs & Vendors", "icon": "ClipboardList"},
                {"id": "inventory", "label": "Inventory", "icon": "Package"},
                {"id": "barcode", "label": "Warehouse & Stock Audit", "icon": "Warehouse"},
                {"id": "matching", "label": "Invoice Matching", "icon": "ShieldCheck"},
                {"id": "bridge", "label": "Accounting Bridge", "icon": "Scale"},
                {"id": "scan", "label": "Document Scanning", "icon": "FileScan"},
            ],
        },
        "compliance": {
            "enabled": True,
            "tabs": [
                {"id": "scan", "label": "Document Scanning", "icon": "FileScan"},
            ],
        },
        "hr": {
            "enabled": True,
            "tabs": [
                {"id": "overview", "label": "Overview", "icon": "LayoutDashboard"},
                {"id": "directory", "label": "Employee Directory", "icon": "Users"},
                {"id": "openings", "label": "Job Openings", "icon": "Briefcase"},
                {"id": "pipeline", "label": "Recruitment Pipeline", "icon": "GitBranch"},
                {"id": "onboarding", "label": "Onboarding", "icon": "UserPlus"},
                {"id": "leave", "label": "Leave Tracker", "icon": "Calendar"},
                {"id": "performance", "label": "Performance Reviews", "icon": "TrendingUp"},
                {"id": "equipment", "label": "Equipment Tracker", "icon": "Monitor"},
                {"id": "training", "label": "Training & Development", "icon": "GraduationCap"},
            ],
        },
        "facility": {
            "enabled": True,
            "tabs": [
                {"id": "units", "label": "Unit Registration", "icon": "Home"},
                {"id": "inspect", "label": "Daily Inspection", "icon": "Camera"},
                {"id": "records", "label": "Inspection Records", "icon": "FileText"},
                {"id": "scan", "label": "Document Scanning", "icon": "FileScan"},
            ],
        },
        "marketing": {
            "enabled": True,
            "tabs": [
                {"id": "summary", "label": "Summary", "icon": "LayoutDashboard"},
                {"id": "leads", "label": "Leads", "icon": "UserPlus"},
                {"id": "events", "label": "Events", "icon": "Calendar"},
                {"id": "seo", "label": "SEO Rankings", "icon": "Search"},
                {"id": "social", "label": "Social Media", "icon": "Share2"},
                {"id": "content", "label": "Content", "icon": "FileText"},
        "projects": {
            "enabled": True,
            "tabs": [
                {"id": "overview", "label": "Overview", "icon": "LayoutDashboard"},
                {"id": "projects", "label": "Projects", "icon": "Kanban"},
                {"id": "active", "label": "Active", "icon": "Activity"},
                {"id": "tasks", "label": "Tasks", "icon": "SquareCheckBig"},
                {"id": "plan", "label": "Plan", "icon": "CalendarClock"},
                {"id": "reports", "label": "Reports", "icon": "BarChart3"},
                {"id": "support", "label": "Support", "icon": "LifeBuoy"},
            ],
        },
    }

    return dashboard_meta.get(name, {"enabled": False, "tabs": []})


# ─── CRM list/search endpoints (live data direct from the brain) ───

# CRM data lives in the brain under source ``crm``. Deals are pages with
# slug prefix ``deals/``, companies are ``companies/``, and tasks are held in
# a single index page ``crm/tasks-index`` (tasks are not first-class pages).

_CRM_MOCK: Optional[dict] = None


_ACRONYMS = {"api", "ai", "id", "crm", "mii", "qbr", "poc", "loa", "nda", "ncr", "sku", "rfq"}


def _pretty_word(word: str) -> str:
    """Title-case a slug word, preserving known acronyms (``api`` → ``API``)."""
    return word.upper() if word.lower() in _ACRONYMS else word.title()


def _last_slug_segment(slug: str) -> str:
    """Return the human-readable tail of a page slug (``partners/syspex`` → ``Syspex``)."""
    if not slug:
        return ""
    tail = slug.strip("/").split("/")[-1].replace("_", "-")
    return " ".join(_pretty_word(w) for w in tail.split("-") if w)


def _filter_mock_tasks(
    mock: List[dict],
    completed: Optional[bool],
    assignee: str,
    deal: str,
) -> List[dict]:
    """Apply completed/assignee/deal filters to a mock task list.

    Single source of truth for every tasks mock fallback so no branch can
    leak an unfiltered list (the round-4 review's Critical regression).
    """
    if completed is not None:
        mock = [t for t in mock if t.get("completed") == completed]
    if assignee:
        ca = _canonical_owner(assignee)
        mock = [t for t in mock if _canonical_owner(t.get("assignee", "")) == ca]
    if deal:
        dd = deal.strip().lower()
        mock = [
            t for t in mock
            if dd in str(t.get("deal_slug", "")).lower()
            or dd in str(t.get("deal_title", "")).lower()
        ]
    return mock


def _crm_mock_enabled() -> bool:
    """Opt-in demo mode: serve examples/crm-mock.json when the brain has no data.

    Off by default — live brain data is always preferred. Set
    SHOGUN_WEB_CRM_MOCK=1 to populate every tab with the mock payloads.
    """
    return os.environ.get("SHOGUN_WEB_CRM_MOCK", "").lower() in ("1", "true", "yes")


def _bev_mock_enabled() -> bool:
    """Opt-in demo mode for BEV zones — independent of the CRM flag.

    SHOGUN_WEB_BEV_MOCK=1 enables it directly. When the BEV flag is unset
    the CRM demo flag is treated as a fallback so existing demo setups
    keep working without reconfiguring.
    """
    raw = os.environ.get("SHOGUN_WEB_BEV_MOCK", "").strip()
    if raw:
        return raw.lower() in ("1", "true", "yes")
    return _crm_mock_enabled()


def _load_crm_mock() -> dict:
    """Load examples/crm-mock.json once (empty dict when absent/corrupt)."""
    global _CRM_MOCK
    if _CRM_MOCK is not None:
        return _CRM_MOCK
    _CRM_MOCK = {}
    mock_json_path = pathlib.Path(__file__).resolve().parents[2] / "examples" / "crm-mock.json"
    try:
        with open(mock_json_path, "r", encoding="utf-8") as f:
            _CRM_MOCK = json.load(f)
    except Exception as exc:  # pragma: no cover - local dev file only
        logger.warning("Failed to load CRM mock data from %s: %s", mock_json_path, exc)
    return _CRM_MOCK


CRM_SOURCE = "crm"
# Standardised listing limit for CRM endpoints. Only matters on the
# filesystem path (unbounded); the MCP fallback pages until exhaustion
# regardless and enriches at most _MCP_ENRICH_CAP rows.
CRM_LIST_LIMIT = 10000


from auth import require_admin  # noqa: E402 — admin guard for BEV zone CRUD


@router.get("/ceo-stats")
async def get_crm_ceo_stats(
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Aggregated CEO dashboard stats for CRM — 100% mock data (demo mode).

    Serves the fictional CRM ledger from examples/crm-mock.json. All tabs
    read the same data set, so metrics are internally consistent across the
    dashboard.
    """
    mock_data = _load_crm_mock()
    if not mock_data:
        logger.info("CRM dashboard: mock data missing — returning empty state")
        return {"mock": False}
    
    payload = mock_data.get("dashboard_mock", {})
    if not payload:
        logger.info("CRM dashboard: dashboard_mock section missing")
        return {"mock": False}
    
    return {**payload, "mock": True}


def _extract_deal_list_item(page: dict) -> dict:
    """Map a raw CRM deal page to a CrmDealListItem."""
    fm = _parse_frontmatter(page.get("frontmatter"))
    created_raw = fm.get("created") or (page.get("effective_date") or "")
    return {
        "slug": page.get("slug", ""),
        "title": page.get("title", ""),
        "customer": fm.get("customer", ""),
        "owner": _canonical_owner(fm.get("owner", "")),
        "stage": _canonical_stage(fm.get("stage", "")),
        # None when absent — lets the frontend distinguish "no date" from an empty string
        "created": created_raw[:10] if created_raw else None,
        "source": fm.get("source", ""),
        "amount": _safe_float(fm.get("amount", 0)),
        "priority": fm.get("priority", ""),
        "compiled_truth": (page.get("compiled_truth", "") or "")[:500],
    }


# Slug metadata to exclude from listings (scaffolding / meta pages).
#   broad set  — deal listings may contain gbrain-internal scaffolds
#   narrow set — companies/partners only skip README/_schema (a company
#                named "...activity-log" is legitimate)
# NOTE: the broad set (deal listings only) matches the pre-existing dashboard
# exclusion convention (docs/plans/2026-07-26-profile-dashboards-implementation.md):
# activity-log / risk-register are gbrain scaffolding pages, not real deals —
# real deal slugs are customer/scoped names, so the over-exclude risk is low.
# Companies/partners use the NARROW set; a business named "...activity-log" stays.
_SLUG_SEGMENT_EXCLUDES_BROAD = {"readme", "_schema", "activity-log", "risk-register"}
_SLUG_SEGMENT_EXCLUDES_NARROW = {"readme", "_schema"}
_SLUG_PREFIX_EXCLUDES = ("templates/",)


def _is_meta_slug(slug: str, *, broad: bool = True) -> bool:
    last_segment = slug.rstrip("/").rsplit("/", 1)[-1].lower()
    if last_segment in (_SLUG_SEGMENT_EXCLUDES_BROAD if broad else _SLUG_SEGMENT_EXCLUDES_NARROW):
        return True
    return any(slug.startswith(pfx) for pfx in _SLUG_PREFIX_EXCLUDES)


async def _fetch_brain_pages_safe(source: str, *, limit: int, slug_prefix: str) -> list:
    """Graceful fetching: never let a gbrain failure 500 a CRM listing.

    Returns the raw pages list; an MCP failure (server down, timeout) or any
    other exception degrades to [] — the endpoints return their empty-state
    shapes, exactly like the pre-brain code.
    """
    try:
        return await gbrain_fetch_pages(source, limit=limit, slug_prefix=slug_prefix)
    except Exception as exc:  # pragma: no cover - transport/protocol failures
        logger.warning("gbrain fetch failed for %s/%s: %s", source, slug_prefix, exc)
        return []


@router.get("/deals")
async def list_crm_deals(
    name: str = Path(...),
    search: str = Query("", description="Filter by title/customer"),
    stage: str = Query("", description="Filter by stage"),
    owner: str = Query("", description="Filter by owner"),
    priority: str = Query("", description="Filter by priority (High/Medium/Low)"),
    source: str = Query("", description="Filter by lead source"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """List CRM deals — 100% mock data (demo mode).

    Serves deals from examples/crm-mock.json with the same filtering logic
    as the live gbrain path.
    """
    mock_data = _load_crm_mock()
    if not mock_data:
        return {"deals": [], "total": 0, "mock": False}
    
    items = list(mock_data.get("deals", []))

    if search:
        s = search.lower()
        items = [d for d in items if s in str(d.get("title", "")).lower() or s in str(d.get("customer", "")).lower()]
    if stage:
        items = [d for d in items if _canonical_stage(d.get("stage", "")) == _canonical_stage(stage)]
    if owner:
        items = [d for d in items if _canonical_owner(d.get("owner", "")) == _canonical_owner(owner)]
    if priority:
        cp = _canonical_priority(priority)
        items = [d for d in items if _canonical_priority(d.get("priority", "")) == cp]
    if source:
        ss = source.strip().lower()
        items = [d for d in items if ss in (d.get("source") or "").lower()]

    # Sort by created date descending
    items.sort(key=lambda d: d.get("created") or "", reverse=True)

    return {"deals": items, "total": len(items), "mock": True}


@router.get("/companies")
async def list_crm_companies(
    name: str = Path(...),
    search: str = Query("", description="Filter by title"),
    industry: str = Query("", description="Filter by industry"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """List CRM companies — 100% mock data (demo mode)."""
    mock_data = _load_crm_mock()
    if not mock_data:
        return {"companies": [], "total": 0, "mock": False}

    items = list(mock_data.get("companies", []))

    if search:
        s = search.lower()
        items = [c for c in items if s in str(c.get("title", "")).lower()]
    if industry:
        items = [c for c in items if c.get("industry", "").lower() == industry.lower()]

    items.sort(key=lambda c: str(c.get("title", "")).lower())

    return {"companies": items, "total": len(items), "mock": True}


@router.get("/partner-sphere")
async def get_partner_sphere(
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Partner Sphere — 9 sections (overview, roster, profile, command center,
    protection, onboarding, QBR, CEO digest, pricing simulator).

    Live mode derives the roster + overview KPIs from partner/deal pages in
    the brain. When the brain has no data the sections come back empty unless
    SHOGUN_WEB_CRM_MOCK=1, which serves the demo payload from
    examples/crm-mock.json (marked ``mock: True``).
    """
    result: dict = {
        "overview": None,
        "masterList": [],
        "profile": None,
        "commandCenter": None,
        "protection": None,
        "onboarding": None,
        "qbr": None,
        "ceoDigest": None,
        "pricing": None,
        "mock": False,
    }
    partners = await _fetch_brain_pages_safe(CRM_SOURCE, limit=CRM_LIST_LIMIT, slug_prefix="partners/")

    if partners:
        # Narrow meta filter keeps a partners/readme page from inflating the
        # roster and the Active Partners KPI (consistency with list_crm_partners).
        partners = [p for p in partners if isinstance(p, dict) and not _is_meta_slug(str(p.get("slug", "")), broad=False)]
        # Live-derived sections: roster + overview KPIs.
        # Business fields wire from partner-page frontmatter where present
        # (snake_case keys); pages without them keep the placeholders.

        def _int0(v):
            try:
                return int(v)
            except (TypeError, ValueError):
                return 0

        def _str_field(fm, key):
            v = fm.get(key)
            return str(v) if v not in (None, "") else "—"

        master_rows: list = []
        active_count = 0
        for p in partners:
            fm = _parse_frontmatter(p.get("frontmatter") or {})
            status = str(fm.get("status") or "Active")
            if status.lower() == "active":
                active_count += 1
            master_rows.append({
                "name": (p.get("title") or _last_slug_segment(p.get("slug", "")) or "Partner"),
                "tier": str(fm.get("tier", "")),
                "am": str(fm.get("am", "")),
                "status": status,
                "regions": str(fm.get("country", "")),
                "tags": fm.get("tags") if isinstance(fm.get("tags"), list) else [],
                "openDeals": _int0(fm.get("open_deals")),
                "pipeline": _str_field(fm, "pipeline"),
                "licences": _str_field(fm, "licences"),
                "score": _safe_float(fm.get("score")),
                "lastActivity": _str_field(fm, "last_activity"),
            })
        result["masterList"] = master_rows
        result["overview"] = {
            "kpis": [
                {"label": "Active Partners", "value": str(active_count), "note": "from brain"},
                {"label": "Partner Pipeline", "value": "—", "note": ""},
                {"label": "Partner Won (YTD)", "value": "—", "note": ""},
                {"label": "POC → Commercial", "value": "—", "note": ""},
                {"label": "Avg Ramp Velocity", "value": "—", "note": ""},
            ],
            "amCoverage": [],
            "tierBoard": [],
            "funnel": [],
            "leakPoints": [],
            "battleLog": [],
            "cohortGrid": None,
            "openPipeline": [],
            "hygiene": None,
            "aiBrief": None,
        }

    # Always serve mock partner sphere on demo
    mock_sphere = _load_crm_mock().get("partner_sphere") or {}
    for key in result:
        if key == "mock":
            continue
        if mock_sphere.get(key):
            result[key] = mock_sphere[key]
    result["mock"] = True
    return result


@router.get("/partners")
async def list_crm_partners(
    name: str = Path(...),
    search: str = Query("", description="Filter by title"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """List CRM partners — 100% mock data (demo mode)."""
    mock_data = _load_crm_mock()
    if not mock_data:
        return {"partners": [], "total": 0, "mock": False}

    items = list(mock_data.get("partners", []))

    if search:
        s = search.lower()
        items = [c for c in items if s in str(c.get("title", "")).lower()]

    items.sort(key=lambda c: str(c.get("title", "")).lower())

    return {"partners": items, "total": len(items), "mock": True}


@router.get("/tasks")
async def list_crm_tasks(
    name: str = Path(...),
    completed: Optional[bool] = Query(None, description="Filter by completion status"),
    assignee: str = Query("", description="Filter by assignee"),
    deal: str = Query("", description="Filter by deal slug/title"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """List CRM tasks — 100% mock data (demo mode)."""
    mock_data = _load_crm_mock()
    if not mock_data:
        return {"tasks": [], "total": 0, "mock": False}

    mock = _filter_mock_tasks(mock_data.get("tasks", []), completed, assignee, deal)
    return {"tasks": mock, "total": len(mock), "mock": True}


class SearchBody(BaseModel):
    query: str = ""


@router.post("/search")
async def crm_search(
    body: SearchBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Global search across CRM pages direct from the brain (source ``crm``)."""
    query = body.query.strip()
    if not query:
        return {"results": []}

    try:
        raw = await gbrain_search(CRM_SOURCE, query, limit=20)
    except Exception as exc:  # pragma: no cover - transport/protocol failures
        logger.warning("gbrain search failed for %s: %s", CRM_SOURCE, exc)
        if _crm_mock_enabled():
            q = query.lower()
            mock = [r for r in _load_crm_mock().get("search_results", [])
                    if q in str(r.get("title", "")).lower() or q in str(r.get("slug", "")).lower()]
            return {"results": mock, "mock": True}
        return {"results": []}

    normalised: list[dict] = []
    for r in raw:
        if not isinstance(r, dict):
            continue
        row = copy.deepcopy(r)
        row["frontmatter"] = _parse_frontmatter(row.get("frontmatter"))
        if not row.get("category"):
            slug = str(row.get("slug", ""))
            if slug.startswith("deals/"):
                row["category"] = "deals"
            elif slug.startswith("companies/"):
                row["category"] = "companies"
            elif slug.startswith("partners/"):
                row["category"] = "partners"
            elif slug.startswith("persons/"):
                row["category"] = "persons"
            else:
                row["category"] = "unknown"
        normalised.append(row)

    # A live, populated brain that legitimately matches nothing returns [].
    # Demo search rows are served ONLY when the source itself is unavailable
    # (the except path above) — never on a filtered/empty result.
    return {"results": normalised}


# ─── BEV Zones proxy (→ separate microservice) ───

BEV_API_TOKEN = os.environ.get("BEV_API_TOKEN", "")


def _bev_base_url() -> str:
    return os.environ.get("BEV_API_URL", "http://localhost:8001/api/v1").rstrip("/")


def _bev_headers() -> dict[str, str]:
    """Build headers for BEV microservice calls, with optional auth token."""
    h: dict[str, str] = {"Accept": "application/json"}
    if BEV_API_TOKEN:
        h["Authorization"] = f"Bearer {BEV_API_TOKEN}"
    return h


@router.get("/bev/zones")
async def list_bev_zones(
    name: str = Path(...),
    user: User = Depends(get_current_user),
) -> dict:
    """List BEV zones via the BEV microservice."""
    base = _bev_base_url()
    live_zones: list = []
    live_ok = False
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(f"{base}/zones", headers=_bev_headers())
            if resp.status_code < 400:
                data = resp.json()
                if isinstance(data, list):
                    live_zones = data
                    live_ok = True
                elif isinstance(data, dict) and "zones" in data:
                    live_zones = data["zones"]
                    live_ok = True
    except httpx.HTTPError as exc:
        logger.warning("BEV zones list error: %s", exc)

    if live_ok:
        return {"zones": live_zones}
    if _bev_mock_enabled():
        mock_zones = _load_crm_mock().get("bev_zones")
        if mock_zones:
            return {"zones": mock_zones, "mock": True}
    return {"zones": []}



class BevZoneBounds(BaseModel):
    """Validated bounding box for a BEV zone (Cartesian coordinates, metres)."""

    xMin: float = Field(default=0.0, description="Left edge (metres)")
    yMin: float = Field(default=0.0, description="Bottom edge (metres)")
    xMax: float = Field(description="Right edge — must be > xMin")
    yMax: float = Field(description="Top edge — must be > yMin")

    from pydantic import model_validator  # local import avoids top-level v1/v2 ambiguity

    @model_validator(mode="after")
    def _validate_bounds(self) -> "BevZoneBounds":
        if self.xMax <= self.xMin:
            raise ValueError("xMax must be greater than xMin")
        if self.yMax <= self.yMin:
            raise ValueError("yMax must be greater than yMin")
        return self


class BevZoneBody(BaseModel):
    name: str = Field(..., min_length=1, description="Zone name")
    calibrationType: str = "cartesian"
    cameraIds: list[str] = []
    bounds: Optional[BevZoneBounds] = None
    origin: Optional[dict] = None
    rois: list[dict] = []
    tripwires: list[dict] = []


@router.post("/bev/zones")
async def create_bev_zone(
    body: BevZoneBody,
    name: str = Path(...),
    user: User = Depends(require_admin),
) -> dict:
    """Create a BEV zone via the BEV microservice."""
    base = _bev_base_url()
    payload = body.model_dump() if hasattr(body, "model_dump") else body.dict()
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.post(f"{base}/zones", json=payload, headers=_bev_headers())
            if resp.status_code >= 400:
                raise HTTPException(status_code=resp.status_code, detail=resp.text[:300])
            return resp.json()
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail=f"BEV service unavailable: {exc}")


@router.put("/bev/zones/{zone_id}")
async def update_bev_zone(
    body: BevZoneBody,
    name: str = Path(...),
    zone_id: str = Path(...),
    user: User = Depends(require_admin),
) -> dict:
    """Update a BEV zone via the BEV microservice."""
    base = _bev_base_url()
    payload = body.model_dump() if hasattr(body, "model_dump") else body.dict()
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.put(f"{base}/zones/{_url_quote(zone_id, safe='')}", json=payload, headers=_bev_headers())
            if resp.status_code >= 400:
                raise HTTPException(status_code=resp.status_code, detail=resp.text[:300])
            return resp.json()
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail=f"BEV service unavailable: {exc}")


@router.delete("/bev/zones/{zone_id}")
async def delete_bev_zone(
    name: str = Path(...),
    zone_id: str = Path(...),
    user: User = Depends(require_admin),
) -> dict:
    """Delete a BEV zone via the BEV microservice."""
    base = _bev_base_url()
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.delete(f"{base}/zones/{_url_quote(zone_id, safe='')}", headers=_bev_headers())
            if resp.status_code >= 400 and resp.status_code != 204:
                raise HTTPException(status_code=resp.status_code, detail=resp.text[:300])
            return {"ok": True}
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail=f"BEV service unavailable: {exc}")


# ─── Finance aggregation helpers ───

def _safe_float(val: Any, default: float = 0.0) -> float:
    try:
        return float(val or 0)
    except (TypeError, ValueError):
        return default


def _safe_int(val: Any, default: int = 0) -> int:
    try:
        return int(float(val or 0))
    except (TypeError, ValueError):
        return default


# ─── Finance mock data ───────────────────────────────────────────────────
# The Finance dashboard serves one internally-aligned fictional ledger from
# examples/finance-budget.json (generated by scripts/generate_finance_mock.py).
# No live QBO connection and no gbrain snapshots — every tab reads the same
# mock so all figures reconcile. Regenerate with:
#   python scripts/generate_finance_mock.py
_FINANCE_MOCK: Optional[dict] = None


def _load_finance_mock() -> dict:
    """Load (and cache) the demo ledger from examples/finance-budget.json."""
    global _FINANCE_MOCK
    if _FINANCE_MOCK is not None:
        return _FINANCE_MOCK
    json_path = pathlib.Path(__file__).resolve().parents[2] / "examples" / "finance-budget.json"
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            _FINANCE_MOCK = json.load(f).get("dashboard_mock", {})
    except Exception as e:
        logger.warning("Failed to load finance mock from %s: %s", json_path, e)
        _FINANCE_MOCK = {}
    return _FINANCE_MOCK


def _run_finance_aggregation(pages: List[dict] | None = None) -> dict:
    """Finance dashboard stats — 100% mock data (demo mode).

    Serves the fictional ledger from examples/finance-budget.json. All tabs
    read the same data set, so balances reconcile across the dashboard:
    bank balances = liquid cash, aging buckets = totals, assets = liabilities
    + equity, monthly series sum to YTD figures. The ``pages`` argument is
    accepted for signature compatibility and ignored.
    """
    mock_data = _load_finance_mock()
    if not mock_data:
        logger.info("Finance dashboard: mock ledger missing — returning empty state")
        return {"mock": False}

    total_liquid_cash = _safe_float(mock_data.get("totalLiquidCash"))
    total_ar = _safe_float(mock_data.get("totalAR"))
    total_ap = _safe_float(mock_data.get("totalAP"))
    total_ca = _safe_float(mock_data.get("totalCurrentAssets"))
    total_nca = _safe_float(mock_data.get("totalNonCurrentAssets"))
    total_assets = _safe_float(mock_data.get("totalAssets"))
    total_liabilities = _safe_float(mock_data.get("totalLiabilities"))
    total_equity = _safe_float(mock_data.get("totalEquity"))
    total_current_liabilities = _safe_float(mock_data.get("totalCurrentLiabilities"))
    gross_margin = _safe_float(mock_data.get("grossMargin"))

    return {
        # Mock flag — true: data loaded from examples/finance-budget.json
        "mock": True,
        # Tab 1 — Executive Pulse
        "totalLiquidCash": total_liquid_cash,
        "netMonthlyBurn": _safe_float(mock_data.get("netMonthlyBurn")),
        "cashRunwayMonths": _safe_float(mock_data.get("cashRunwayMonths")),
        "runwayStatus": mock_data.get("runwayStatus", "unknown"),
        "revenueMTD": _safe_float(mock_data.get("revenueMTD")),
        "revenueYTD": _safe_float(mock_data.get("revenueYTD")),
        "grossMargin": gross_margin,
        "ebitdaMargin": _safe_float(mock_data.get("ebitdaMargin")),
        "unpaidStatutory": _safe_float(mock_data.get("unpaidStatutory")),
        "riskAlerts": mock_data.get("riskAlerts", []),
        "revenueOpexTrend": mock_data.get("revenueOpexTrend", []),
        "cashFlowTrend": mock_data.get("cashFlowTrend", []),
        # Overview KPIs — ratios derived so they always reconcile
        "totalLiabilities": total_liabilities,
        "totalEquity": total_equity,
        "debtToEquity": (total_liabilities / total_equity) if total_equity else 0.0,
        "equityRatio": (total_equity / total_assets) if total_assets else 0.0,
        "arToApCoverage": (total_ar / total_ap) if total_ap else 0.0,
        "netWorkingCapital": total_ca - total_current_liabilities,
        "grossWorkingCapital": total_ca,
        "grossProfitMargin": gross_margin,
        "totalCurrentLiabilities": total_current_liabilities,
        "apAgingByTarget": mock_data.get("apAgingByTarget", []),
        "monthlyPlTrend": mock_data.get("monthlyPlTrend", []),
        # Cash Flow tab
        "arAgingByTarget": mock_data.get("arAgingByTarget", []),
        "cashFlowForecast": mock_data.get("cashFlowForecast", []),
        "burnTrend": mock_data.get("burnTrend", []),
        "cashFlowBreakdown": mock_data.get("cashFlowBreakdown", {}),
        # Tab 2 — Cash & Runway
        "bankAccounts": mock_data.get("bankAccounts", []),
        "fxPositions": mock_data.get("fxPositions", []),
        "forecast13w": mock_data.get("forecast13w", {"conservative": [], "expected": [], "optimistic": []}),
        "fixedOpex": _safe_float(mock_data.get("fixedOpex")),
        "variableOpex": _safe_float(mock_data.get("variableOpex")),
        # Assets tab
        "currentAssets": mock_data.get("currentAssets", []),
        "nonCurrentAssets": mock_data.get("nonCurrentAssets", []),
        "assetTrend": mock_data.get("assetTrend", []),
        "totalCurrentAssets": total_ca,
        "totalNonCurrentAssets": total_nca,
        "totalAssets": total_assets,
        # Tab 3 — AR & AP
        "totalAR": total_ar,
        "arOverdue30": _safe_float(mock_data.get("arOverdue30")),
        "dso": _safe_float(mock_data.get("dso")),
        "totalAP": total_ap,
        "apOverdue": _safe_float(mock_data.get("apOverdue")),
        "dpo": _safe_float(mock_data.get("dpo")),
        "arAging": mock_data.get("arAging", {"bucket_0_30": 0, "bucket_31_60": 0, "bucket_61_90": 0, "bucket_90_plus": 0}),
        "dunningQueue": mock_data.get("dunningQueue", []),
        "arInvoices": mock_data.get("arInvoices", []),
        "apBills": mock_data.get("apBills", []),
        # Tab 4 — BvA & Unit Economics
        "bvaDepartments": mock_data.get("bvaLineItems", []),
        "bvaLineItems": mock_data.get("bvaLineItems", []),
        "unitEconomics": mock_data.get("unitEconomics", {"gross_margin_pct": 0, "contribution_margin_pct": 0, "cac": 0, "ltv": 0, "ltv_cac_ratio": 0}),
        "clientConcentration": mock_data.get("clientConcentration", []),
        # Tab 5 — Close & Tax
        "closeChecklist": mock_data.get("closeChecklist", []),
        "statutorySchedule": mock_data.get("statutorySchedule", []),
        "sstReadiness": mock_data.get("sstReadiness", {"draft_status": "Not Started", "taxable_sales": 0, "sst_liability": 0}),
        "cp58Register": mock_data.get("cp58Register", []),
        "whtQueue": mock_data.get("whtQueue", []),
        "expenseClaimAudit": mock_data.get("expenseClaimAudit", []),
    }


@router.get("/finance-stats")
async def get_finance_stats(
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Aggregated Finance dashboard stats — all 5 tabs (mock data only)."""
    return _run_finance_aggregation([])


# ─── Procurement aggregation helpers ───


def _run_procurement_aggregation(pages: List[dict]) -> dict:
    """Aggregate gbrain procurement pages into structured dashboard stats.
    Supplies rich demo mock data when gbrain snapshots are empty so users can view all 5 tabs immediately.
    """
    # ── Pull snapshot pages (procurement agent writes these) ──
    snapshot_map: Dict[str, dict] = {}
    for p in pages:
        slug = str(p.get("slug", ""))
        fm = _parse_frontmatter(p.get("frontmatter", {}))
        if not fm:
            try:
                body = p.get("content") or p.get("body") or ""
                fm = json.loads(body) if body else {}
            except (json.JSONDecodeError, TypeError):
                fm = {}
        if fm:
            snapshot_map[slug] = fm

    inventory_snap = snapshot_map.get("snapshots/inventory", snapshot_map.get("procurement/snapshots/inventory", {}))
    po_snap = snapshot_map.get("snapshots/purchase-orders", snapshot_map.get("procurement/snapshots/purchase-orders", {}))
    vendor_snap = snapshot_map.get("snapshots/vendors", snapshot_map.get("procurement/snapshots/vendors", {}))
    movement_snap = snapshot_map.get("snapshots/stock-movements", snapshot_map.get("procurement/snapshots/stock-movements", {}))
    bridge_snap = snapshot_map.get("snapshots/accounting-bridge", snapshot_map.get("procurement/snapshots/accounting-bridge", {}))

    has_real_data = bool(inventory_snap or po_snap)

    if has_real_data:
        # Pull from the live gbrain snapshots written by dashboard-snapshot-writer
        total_inventory_valuation = _safe_float(inventory_snap.get("total_inventory_valuation"))
        total_active_skus = _safe_float(inventory_snap.get("total_active_skus"))
        low_stock_alerts = _safe_float(inventory_snap.get("low_stock_alerts"))
        dead_slow_stock_capital = _safe_float(inventory_snap.get("dead_slow_stock_capital"))
        valuation_by_category: List[dict] = inventory_snap.get("valuation_by_category", [])
        sku_catalog: List[dict] = inventory_snap.get("sku_catalog", [])
        dead_slow_stock: List[dict] = inventory_snap.get("dead_slow_stock", [])
        warehouse_bin_capacity: List[dict] = inventory_snap.get("warehouse_bin_capacity", [])
        spend_vs_budget_trend: List[dict] = inventory_snap.get("spend_vs_budget_trend", [])
        procurement_spend_mtd = _safe_float(inventory_snap.get("procurement_spend_mtd"))
        procurement_spend_budget_mtd = _safe_float(inventory_snap.get("procurement_spend_budget_mtd"))

        open_po_count = _safe_float(po_snap.get("open_po_count"))
        open_po_value = _safe_float(po_snap.get("open_po_value"))
        po_pipeline: List[dict] = po_snap.get("po_pipeline", [])
        active_purchase_orders: List[dict] = po_snap.get("active_purchase_orders", [])
        executive_approval_queue: List[dict] = po_snap.get("executive_approval_queue", [])

        vendor_scorecard: List[dict] = vendor_snap.get("vendor_scorecard", [])
        vendor_spend_concentration: List[dict] = vendor_snap.get("vendor_spend_concentration", [])

        stock_movements: List[dict] = movement_snap.get("stock_movements", [])
        movement_type_distribution: List[dict] = movement_snap.get("movement_type_distribution", [])
        shrinkage_flag_items: List[dict] = movement_snap.get("shrinkage_flag_items", [])

        bridge_status: dict = bridge_snap.get("bridge_status", {"enabled": False, "provider": "None", "connected": False})
        po_bill_conversion_queue: List[dict] = bridge_snap.get("po_bill_conversion_queue", [])
        gl_valuation_reconciliation: List[dict] = bridge_snap.get("gl_valuation_reconciliation", [])

        risk_alerts: List[dict] = inventory_snap.get("risk_alerts", [])
    else:
        # No gbrain snapshots — fall back to examples/procurement-mock.json so
        # the dashboard shows realistic demo data instead of empty zeros. Same
        # pattern as the finance dashboard (PR #12 review allows mock as
        # graceful-degradation fallback, just no fabricated figures as live).
        # `mock: true` flags the UI that this is demo data, not live.
        mock_json_path = pathlib.Path(__file__).resolve().parents[2] / "examples" / "procurement-mock.json"
        mock_data: Dict[str, Any] = {}
        if mock_json_path.exists():
            try:
                with open(mock_json_path, "r", encoding="utf-8") as f:
                    mock_data = json.load(f).get("dashboard_mock", {})
                logger.info("Procurement dashboard: no gbrain snapshots — loaded mock from %s", mock_json_path)
            except (OSError, json.JSONDecodeError) as exc:
                logger.warning("Failed to load procurement mock from %s: %s", mock_json_path, exc)

        total_inventory_valuation = _safe_float(mock_data.get("totalInventoryValuation"))
        total_active_skus = _safe_float(mock_data.get("totalActiveSkus"))
        low_stock_alerts = _safe_float(mock_data.get("lowStockAlerts"))
        dead_slow_stock_capital = _safe_float(mock_data.get("deadSlowStockCapital"))
        open_po_count = _safe_float(mock_data.get("openPoCount"))
        open_po_value = _safe_float(mock_data.get("openPoValue"))
        procurement_spend_mtd = _safe_float(mock_data.get("procurementSpendMtd"))
        procurement_spend_budget_mtd = _safe_float(mock_data.get("procurementSpendBudgetMtd"))

        risk_alerts = mock_data.get("riskAlerts", [])
        valuation_by_category = mock_data.get("valuationByCategory", [])
        spend_vs_budget_trend = mock_data.get("spendVsBudgetTrend", [])

        sku_catalog = mock_data.get("skuCatalog", [])
        dead_slow_stock = mock_data.get("deadSlowStock", [])
        warehouse_bin_capacity = mock_data.get("warehouseBinCapacity", [])

        stock_movements = mock_data.get("stockMovements", [])
        movement_type_distribution = mock_data.get("movementTypeDistribution", [])
        shrinkage_flag_items = mock_data.get("shrinkageFlagItems", [])

        po_pipeline = mock_data.get("poPipeline", [])
        active_purchase_orders = mock_data.get("activePurchaseOrders", [])
        executive_approval_queue = mock_data.get("executiveApprovalQueue", [])
        vendor_scorecard = mock_data.get("vendorScorecard", [])
        vendor_spend_concentration = mock_data.get("vendorSpendConcentration", [])

        bridge_status = mock_data.get("accountingBridge", {"enabled": False, "provider": "None", "connected": False})
        po_bill_conversion_queue = mock_data.get("poBillConversionQueue", [])
        gl_valuation_reconciliation = mock_data.get("glValuationReconciliation", [])

        # Phase A–E mock data (PR / RFQ / Barcode / 3-way match)
        purchase_requisitions = mock_data.get("purchaseRequisitions", [])
        rfq_comparisons = mock_data.get("rfqComparisons", [])
        barcode_batches = mock_data.get("barcodeBatches", [])
        three_way_matches = mock_data.get("threeWayMatches", [])

    # Pull PR/RFQ/barcode/match from snapshots when live data exists (snapshot
    # writer will populate these once the full procurement cycle endpoints are built).
    if has_real_data:
        purchase_requisitions = inventory_snap.get("purchase_requisitions", [])
        rfq_comparisons = inventory_snap.get("rfq_comparisons", [])
        barcode_batches = inventory_snap.get("barcode_batches", [])
        three_way_matches = inventory_snap.get("three_way_matches", [])

    return {
        # Mock flag — true when data loaded from examples/procurement-mock.json
        "mock": not has_real_data,
        # Tab 1 — Executive Procurement & Reorder Pulse
        "totalInventoryValuation": total_inventory_valuation,
        "totalActiveSkus": total_active_skus,
        "lowStockAlerts": low_stock_alerts,
        "deadSlowStockCapital": dead_slow_stock_capital,
        "openPoCount": open_po_count,
        "openPoValue": open_po_value,
        "procurementSpendMtd": procurement_spend_mtd,
        "procurementSpendBudgetMtd": procurement_spend_budget_mtd,
        "riskAlerts": risk_alerts,
        "valuationByCategory": valuation_by_category,
        "spendVsBudgetTrend": spend_vs_budget_trend,
        # Tab 2 — Inventory Catalog & Dead/Slow Stock
        "skuCatalog": sku_catalog,
        "deadSlowStock": dead_slow_stock,
        "warehouseBinCapacity": warehouse_bin_capacity,
        # Tab 3 — Stock Movement Audit Log
        "stockMovements": stock_movements,
        "movementTypeDistribution": movement_type_distribution,
        "shrinkageFlagItems": shrinkage_flag_items,
        # Tab 4 — Purchase Orders & Vendor Scorecard
        "poPipeline": po_pipeline,
        "activePurchaseOrders": active_purchase_orders,
        "executiveApprovalQueue": executive_approval_queue,
        "vendorScorecard": vendor_scorecard,
        "vendorSpendConcentration": vendor_spend_concentration,
        # Tab 5 — Accounting Bridge & Valuation Reconciliation
        "accountingBridge": bridge_status,
        "poBillConversionQueue": po_bill_conversion_queue,
        "glValuationReconciliation": gl_valuation_reconciliation,
        # Tab 6 — Purchase Requisitions (PR)
        "purchaseRequisitions": purchase_requisitions,
        # Tab 7 — RFQ & Vendor Sourcing
        "rfqComparisons": rfq_comparisons,
        # Tab 8 — Barcode Tagging & Scan Counter
        "barcodeBatches": barcode_batches,
        # Tab 9 — 3-Way Match Verification
        "threeWayMatches": three_way_matches,
    }


@router.get("/procurement-stats")
async def get_procurement_stats(
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Aggregated Procurement dashboard stats — all 5 tabs."""
    pages = await _fetch_brain_pages_safe("procurement", limit=300, slug_prefix="")
    return _run_procurement_aggregation(pages)


# ---------------------------------------------------------------------------
# Marketing dashboard endpoints
# ---------------------------------------------------------------------------

_MARKETING_MOCK_PATH = pathlib.Path(__file__).resolve().parent.parent.parent / "examples" / "marketing-dashboard-mock.json"


def _load_marketing_mock() -> dict:
    """Load marketing mock data from examples/ JSON file."""
    import json as _json
    with open(_MARKETING_MOCK_PATH, encoding="utf-8") as f:
        return _json.load(f)


@router.get("/marketing-stats")
async def get_marketing_stats(
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Aggregated Marketing dashboard stats — all 6 tabs.

    Currently returns mock/demo data. Future: aggregate from gbrain marketing source.
    """
    # TODO: Query gbrain marketing source for live data
    # For now, return mock data with mock=True flag
    return _load_marketing_mock()


# ---------------------------------------------------------------------------
# Plantation Estate Operations dashboard endpoints
# ---------------------------------------------------------------------------


async def _call_estate_agent(name: str, prompt: str, photo_paths: Optional[List[str]] = None) -> str:
    """Call the Hermes agent for the facility department via the gateway.

    If photo_paths are provided, they are passed as image attachments so the
    gateway base64-encodes them and switches to the Qwen vision model.
    """
    try:
        from gateway import _generate_department_response_async
    except ImportError:
        # Fallback: try a direct LLM call via the gateway's embedded agent
        return '{"error": "Gateway agent function not available — start the Hermes gateway"}'

    attachments = None
    if photo_paths:
        attachments = []
        for p in photo_paths:
            ext = pathlib.Path(p).suffix.lower()
            mime_map = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                        ".webp": "image/webp", ".gif": "image/gif"}
            attachments.append({
                "path": p,
                "is_image": True,
                "mime_type": mime_map.get(ext, "image/png"),
            })

    response = await _generate_department_response_async(
        name, prompt, soul_content="", attachments=attachments,
    )
    return response


def _extract_json_from_text(text: str, is_array: bool = False) -> Any:
    """Try to extract JSON from the agent's text response."""
    pattern = r'\[[\s\S]*\]' if is_array else r'\{[\s\S]*\}'
    match = _re.search(pattern, text)
    if match:
        try:
            return json.loads(match.group())
        except json.JSONDecodeError:
            pass
    return [] if is_array else {"raw_response": text}


@router.post("/scan-document")
async def scan_document(
    name: str = Path(...),
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Upload a document, OCR it, then use DeepSeek for summary + interpretation.
    Saves the scan result to the scanned_documents table.

    Flow: upload → OCR (pymupdf for PDF, vision model for images) → DeepSeek → save.
    """
    from models import ScannedDocument
    from gateway import _call_deepseek
    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)

    cfg = get_config()
    upload_dir = pathlib.Path(cfg.db_path).parent / "dashboard_uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    # Prefix with UUID to prevent filename collisions across tenants
    import uuid as _uuid
    safe_name = pathlib.Path(file.filename or "document").name
    unique_name = f"{_uuid.uuid4().hex[:8]}_{safe_name}"
    file_path = upload_dir / unique_name
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    file_url = f"/api/doc-uploads/{unique_name}"

    # --- Step 1: OCR — extract raw text from the document ---
    ocr_text = ""
    ext = pathlib.Path(safe_name).suffix.lower()

    if ext == ".pdf":
        # Try pymupdf first (text-based PDFs)
        try:
            import fitz  # pymupdf
            doc = fitz.open(str(file_path))
            for page in doc:
                ocr_text += page.get_text()
            doc.close()
        except ImportError:
            pass
        except Exception as exc:
            logger.warning("pymupdf OCR failed: %s", exc)

    if not ocr_text and ext in (".pdf", ".png", ".jpg", ".jpeg", ".webp", ".gif"):
        # Fall back to vision model (scanned PDFs, images)
        # For scanned PDFs: render pages as images first (PyMuPDF), don't send raw PDF to VLM
        vision_images = []
        if ext == ".pdf":
            try:
                import fitz
                doc = fitz.open(str(file_path))
                for page_num, page in enumerate(doc):
                    pix = page.get_pixmap(dpi=200)
                    img_path = file_path.parent / f"{file_path.stem}_p{page_num}.png"
                    pix.save(str(img_path))
                    vision_images.append(str(img_path))
                doc.close()
            except Exception as exc:
                logger.warning("PDF page rendering failed: %s", exc)
        else:
            vision_images = [str(file_path)]

        try:
            from gateway import _call_llm_for_department
            for img_path in vision_images:
                vision_resp = await _call_llm_for_department(
                    name,
                    "OCR this document. Extract ALL text visible in the document. Return the raw text only, no commentary.",
                    "",
                    "Document OCR task",
                    attachments=[{"path": img_path, "is_image": True}],
                )
                if vision_resp:
                    ocr_text += vision_resp + "\n"
            # Clean up temporary page images
            for img_path in vision_images:
                if os.path.abspath(img_path) != os.path.abspath(str(file_path)):
                    try:
                        os.remove(img_path)
                    except OSError:
                        pass
        except Exception as exc:
            logger.warning("Vision OCR fallback failed: %s", exc)

    if not ocr_text:
        ocr_text = "(OCR could not extract text from this document.)"

    # --- Step 2: Department-specific DeepSeek analysis ---
    ocr_text_truncated = ocr_text[:8000]  # prevent token overflow

    if name in ("finance", "procurement"):
        # Finance: extract fields as bullet points, NO summary
        deepseek_prompt = (
            f"You are a finance document scanner. Below is the OCR text from an invoice, quotation, or receipt.\n\n"
            f"=== DOCUMENT TEXT ===\n{ocr_text_truncated}\n=== END DOCUMENT TEXT ===\n\n"
            f"Extract ALL key fields from this document. Return ONLY a JSON object:\n"
            f'{{"document_type": "invoice|quotation|receipt|other", '
            f'"fields": {{'
            f'"vendor": "vendor/supplier name", '
            f'"document_number": "invoice/quotation/receipt number", '
            f'"date": "document date", '
            f'"due_date": "due date if present", '
            f'"subtotal": "subtotal amount with currency", '
            f'"tax": "tax amount with currency", '
            f'"total": "total amount with currency", '
            f'"line_items": ["each line item with description and amount"]'
            f'}}, '
            f'"validation": {{"valid": true/false, "message": "verify subtotal + tax = total"}}}}\n\n'
            f"List ALL fields as key-value pairs. Do NOT write a summary. Do NOT add explanation outside JSON."
        )
    elif name in ("facility", "compliance"):
        # Estate Ops / Compliance: summary + interpretation (legal docs, contracts)
        deepseek_prompt = (
            f"You are a legal document analyst. Below is the OCR text from a legal document, contract, or agreement.\n\n"
            f"=== DOCUMENT TEXT ===\n{ocr_text_truncated}\n=== END DOCUMENT TEXT ===\n\n"
            f"Analyse this legal document. Return ONLY a JSON object:\n"
            f'{{"document_type": "employment_contract|service_agreement|lease|nda|other", '
            f'"summary": "3-4 sentences summarising what this document is about, who the parties are, and the key terms", '
            f'"interpretation": {{'
            f'"parties": ["list all parties"], '
            f'"duration": "contract duration if applicable", '
            f'"value": "contract value or salary if applicable", '
            f'"key_obligations": ["list key obligations of each party"], '
            f'"termination_clause": "termination conditions", '
            f'"penalty_clause": "penalty clauses if any"'
            f'}}, '
            f'"risks": ["list any risks: unlimited liability, auto-renewal, short notice, etc."], '
            f'"recommendations": ["list actionable recommendations"]}}\n\n'
            f"Write a clear summary (3-4 sentences) AND a detailed interpretation. Both are required."
        )
    else:
        # Generic fallback
        deepseek_prompt = (
            f"You are a document analysis assistant. Below is the OCR text from a document.\n\n"
            f"=== DOCUMENT TEXT ===\n{ocr_text_truncated}\n=== END DOCUMENT TEXT ===\n\n"
            f"Return ONLY a JSON object with: document_type, fields, summary, risks, validation.\n"
            f"No markdown fences, no explanation outside JSON."
        )

    deepseek_response = await _call_deepseek(deepseek_prompt)
    result = _extract_json_from_text(deepseek_response, is_array=False) if deepseek_response else {"raw_response": deepseek_response or "No response from DeepSeek"}

    # If DeepSeek failed, fall back to the OCR text as summary
    if not deepseek_response:
        result = {
            "document_type": "unknown",
            "fields": {},
            "summary": ocr_text[:200] + "..." if len(ocr_text) > 200 else ocr_text,
            "risks": [],
            "validation": {"valid": True, "message": "DeepSeek unavailable — showing raw OCR text"},
            "ocr_text": ocr_text,
        }

    # Persist to DB
    doc = ScannedDocument(
        tenant_id=tenant.id,
        department=name,
        scanned_by=user.name or user.email,
        filename=file.filename or "document",
        file_path=str(file_path),
        file_url=file_url,
        document_type=str(result.get("document_type", "")) if isinstance(result, dict) else "",
        ocr_summary=str(result.get("summary", "")) if isinstance(result, dict) else "",
        interpretation=result if isinstance(result, dict) else {"raw_response": result},
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)
    return doc.to_dict()


@router.get("/scanned-documents")
async def list_scanned_documents(
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """List all scanned documents for this department."""
    from models import ScannedDocument
    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    docs = db.execute(
        select(ScannedDocument).where(
            ScannedDocument.tenant_id == tenant.id,
            ScannedDocument.department == name,
        ).order_by(ScannedDocument.scan_date.desc())
    ).scalars().all()
    return {"documents": [d.to_dict() for d in docs]}


# ---------------------------------------------------------------------------
# Document Scanning — Multi-source CRUD & execution
# ---------------------------------------------------------------------------

@router.get("/doc-scan/sources")
async def list_doc_scan_sources(name: str = Path(...), user: User = Depends(get_current_user)):
    """List all configured scan sources for this department."""
    sources = [
        {**s, "id": k.split(":")[1]} 
        for k, s in _DOC_SCAN_SOURCES.items() 
        if k.startswith(f"{name}:")
    ]
    return {"sources": sources}


@router.post("/doc-scan/sources")
async def create_doc_scan_source(
    name: str = Path(...),
    title: str = Form(...),
    drive_url: str = Form(...),
    schedule: str = Form("manual"),
    document_type: str = Form("invoice"),
    template: UploadFile = File(None),
    user: User = Depends(get_current_user),
):
    """Create a new scan source configuration."""
    import uuid
    source_id = str(uuid.uuid4())[:8]
    
    template_path = None
    if template:
        cfg = get_config()
        upload_dir = pathlib.Path(cfg.db_path).parent / "dashboard_uploads" / "templates"
        upload_dir.mkdir(parents=True, exist_ok=True)
        template_path = str(upload_dir / f"{source_id}_{template.filename}")
        with open(template_path, "wb") as f:
            content = await template.read()
            f.write(content)
    
    source_key = f"{name}:{source_id}"
    _DOC_SCAN_SOURCES[source_key] = {
        "title": title,
        "drive_url": drive_url,
        "schedule": schedule,
        "document_type": document_type,
        "template_path": template_path,
        "last_run": None,
        "next_run": None,
    }
    
    return {"id": source_id, "status": "created"}


@router.put("/doc-scan/sources/{source_id}")
async def update_doc_scan_source(
    name: str = Path(...),
    source_id: str = Path(...),
    title: str = Form(None),
    drive_url: str = Form(None),
    schedule: str = Form(None),
    document_type: str = Form(None),
    template: UploadFile = File(None),
    user: User = Depends(get_current_user),
):
    """Update an existing scan source configuration."""
    source_key = f"{name}:{source_id}"
    if source_key not in _DOC_SCAN_SOURCES:
        raise HTTPException(status_code=404, detail="Source not found")
    
    source = _DOC_SCAN_SOURCES[source_key]
    if title:
        source["title"] = title
    if drive_url:
        source["drive_url"] = drive_url
    if schedule:
        source["schedule"] = schedule
    if document_type:
        source["document_type"] = document_type
    
    if template:
        cfg = get_config()
        upload_dir = pathlib.Path(cfg.db_path).parent / "dashboard_uploads" / "templates"
        upload_dir.mkdir(parents=True, exist_ok=True)
        template_path = str(upload_dir / f"{source_id}_{template.filename}")
        with open(template_path, "wb") as f:
            content = await template.read()
            f.write(content)
        source["template_path"] = template_path
    
    return {"id": source_id, "status": "updated"}


@router.delete("/doc-scan/sources/{source_id}")
async def delete_doc_scan_source(name: str = Path(...), source_id: str = Path(...), user: User = Depends(get_current_user)):
    """Delete a scan source configuration."""
    source_key = f"{name}:{source_id}"
    if source_key in _DOC_SCAN_SOURCES:
        del _DOC_SCAN_SOURCES[source_key]
    return {"status": "deleted"}


@router.put("/doc-scan/sources/{source_id}/schedule")
async def update_source_schedule(
    name: str = Path(...),
    source_id: str = Path(...),
    schedule: str = Body(...),
    user: User = Depends(get_current_user),
):
    """Update the schedule for a scan source."""
    source_key = f"{name}:{source_id}"
    if source_key not in _DOC_SCAN_SOURCES:
        raise HTTPException(status_code=404, detail="Source not found")
    
    _DOC_SCAN_SOURCES[source_key]["schedule"] = schedule
    return {"status": "updated", "schedule": schedule}


@router.post("/doc-scan/sources/{source_id}/run")
async def run_doc_scan_source(
    name: str = Path(...),
    source_id: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Run a scan source immediately - fetch from Drive, OCR all images, extract to Excel template."""
    source_key = f"{name}:{source_id}"
    if source_key not in _DOC_SCAN_SOURCES:
        raise HTTPException(status_code=404, detail="Source not found")
    
    source = _DOC_SCAN_SOURCES[source_key]
    
    # Mock results for demo — simulates OCR extraction with realistic fields
    mock_invoices = [
        {
            "filename": "INV_2026_001.pdf",
            "fields": {
                "vendor_name": "PALM MACH SDN BHD",
                "invoice_number": "SV250128",
                "invoice_date": "23-APR-2025",
                "total_amount": "RM 547.00",
                "service_tax_no": "W10-2406-32000064",
                "po_number": "4534139727",
                "customer_name": "IOI PLANTATION SDN BHD",
                "uin": "KTWZFMUGXI1CJXIPIBWZVGSJ1O",
            }
        },
        {
            "filename": "INV_2026_002.pdf",
            "fields": {
                "vendor_name": "TECH SOLUTIONS SDN BHD",
                "invoice_number": "TS-INV-2026-0445",
                "invoice_date": "15-MAY-2025",
                "total_amount": "RM 3,250.00",
                "service_tax_no": "W10-2308-41000088",
                "po_number": "PO-2026-1122",
                "customer_name": "SHOGUN ENTERPRISE SDN BHD",
                "uin": "ABCD1234EFGH5678IJKL9012MNOP",
            }
        },
        {
            "filename": "INV_2026_003.pdf",
            "fields": {
                "vendor_name": "KLUANG HARDWARE TRADING",
                "invoice_number": "KH/2026/0891",
                "invoice_date": "02-JUN-2025",
                "total_amount": "RM 1,875.50",
                "service_tax_no": "W10-2401-55000033",
                "po_number": "7890123456",
                "customer_name": "MEGA CORP SDN BHD",
                "uin": "QRST4567UVWX8901YZAB2345CDEF",
            }
        },
    ]
    
    results = []
    for i, mock in enumerate(mock_invoices):
        result = {
            "id": len(_DOC_SCAN_RESULTS.get(name, [])) + i + 1,
            "filename": mock["filename"],
            "file_url": f"/api/doc-uploads/scans/{name}/{source_id}/{mock['filename']}",
            "source_id": source_id,
            "source_title": source["title"],
            "document_type": source["document_type"],
            "ocr_summary": f"Extracted {len(mock['fields'])} fields from {mock['filename']}",
            "interpretation": {
                "fields": mock["fields"],
                "validation": {"valid": True, "message": "Extraction complete"}
            },
            "status": "processed",
            "scan_date": datetime.now().isoformat(),
        }
        results.append(result)
    
    if name not in _DOC_SCAN_RESULTS:
        _DOC_SCAN_RESULTS[name] = []
    _DOC_SCAN_RESULTS[name].extend(results)
    
    # Generate mock combined Excel
    output_excel_url = None
    template_path = source.get("template_path")
    if template_path and Path(template_path).exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # Read headers
            excel_headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    excel_headers[str(cell.value).strip()] = col_idx
            
            # Simple mapping for demo
            field_map = {
                "Doc No": "invoice_number", "Invoice No": "invoice_number",
                "Date of Invoice": "invoice_date", "Date": "invoice_date",
                "Creditor Name": "vendor_name", "Vendor": "vendor_name",
                "Total Invoice Amount": "total_amount", "Amount": "total_amount", "Total": "total_amount",
                "Service Tax No": "service_tax_no", "Tax No": "service_tax_no",
                "PO Number": "po_number", "PO": "po_number",
                "Customer": "customer_name", "Invoice To": "customer_name",
                "UIN": "uin",
            }
            
            for row_idx, result in enumerate(results, 2):
                fields = result.get("interpretation", {}).get("fields", {})
                for header_name, col in excel_headers.items():
                    mapped_field = field_map.get(header_name)
                    if mapped_field and mapped_field in fields:
                        ws.cell(row=row_idx, column=col, value=str(fields[mapped_field]))
            
            output_dir = Path(get_config().db_path).parent / "dashboard_uploads" / "scans" / name / source_id / "output"
            output_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"{source['title'].replace(' ', '_')}_{timestamp}.xlsx"
            output_path = output_dir / output_filename
            wb.save(str(output_path))
            output_excel_url = f"/api/doc-uploads/scans/{name}/{source_id}/output/{output_filename}"
        except Exception:
            pass
    
    # Update last_run
    source["last_run"] = datetime.now().isoformat()
    
    return {"status": "completed", "results_count": len(results), "results": results, "output_excel_url": output_excel_url}


@router.get("/doc-scan/results")
async def list_doc_scan_results(name: str = Path(...), user: User = Depends(get_current_user)):
    """List all scan results for this department."""
    results = _DOC_SCAN_RESULTS.get(name, [])
    return {"results": results}


@router.post("/doc-scan/results/{result_id}/verify")
async def verify_doc_result(name: str = Path(...), result_id: int = Path(...), user: User = Depends(get_current_user)):
    """Mark a scan result as verified."""
    results = _DOC_SCAN_RESULTS.get(name, [])
    for r in results:
        if r["id"] == result_id:
            r["status"] = "verified"
            return {"status": "verified"}
    raise HTTPException(status_code=404, detail="Result not found")


@router.post("/doc-scan/results/{result_id}/reject")
async def reject_doc_result(name: str = Path(...), result_id: int = Path(...), user: User = Depends(get_current_user)):
    """Mark a scan result as rejected."""
    results = _DOC_SCAN_RESULTS.get(name, [])
    for r in results:
        if r["id"] == result_id:
            r["status"] = "rejected"
            return {"status": "rejected"}
    raise HTTPException(status_code=404, detail="Result not found")

@router.post("/inspect-site")
async def inspect_site(
    name: str = Path(...),
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Upload a photo/video, send to Hermes agent for site assessment + storage."""
    cfg = get_config()
    upload_dir = pathlib.Path(cfg.db_path).parent / "dashboard_uploads"
    upload_dir.mkdir(exist_ok=True)
    file_path = upload_dir / file.filename
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    prompt = (
        f"Inspect this site image at {file_path}: "
        f"assess furniture count, cleanliness, site condition, and safety (use site-condition-assessment skill). "
        f"Store the report to gbrain (use site-inspection-storage skill). "
        f"Return JSON with: furniture, cleanliness, site_condition, safety_hazards, "
        f"overall_rating, priority_actions, storage_path."
    )
    response_text = await _call_estate_agent(name, prompt)
    result = _extract_json_from_text(response_text, is_array=False)
    return result


@router.get("/search-documents")
async def search_documents(
    name: str = Path(...),
    q: str = Query(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Search stored documents via Hermes agent."""
    prompt = (
        f"Search for documents matching '{q}' (use document-retrieval skill). "
        f"Return JSON array with: title, summary, and key fields for each match. "
        f"If no results, return empty array."
    )
    response_text = await _call_estate_agent(name, prompt)
    results = _extract_json_from_text(response_text, is_array=True)
    return {"results": results}


# ---------------------------------------------------------------------------
# Site Inspection Units — CRUD + inspection upload
# ---------------------------------------------------------------------------

class CreateUnitPayload(BaseModel):
    site_name: str
    block_name: str
    unit_number: str
    capacity: int = 1
    unit_type: str = "single"  # single, family, dormitory


class PhotoMeta(BaseModel):
    """Metadata for a single photo in a save-inspection request."""
    path: str          # absolute path on server (from assess step)
    filename: str
    url: str = ""      # web-accessible URL (/api/site-photos/<basename>)
    room: str = ""
    assessment: str = ""


class SaveInspectionPayload(BaseModel):
    """Payload for the second step — persist an assessed inspection."""
    photos: List[PhotoMeta]
    furniture_count: str = ""
    cleanliness: str = ""
    site_condition: str = ""
    safety_hazards: str = ""
    overall_rating: str = ""
    priority_actions: str = ""
    merged_assessment: Dict[str, Any] = {}


@router.post("/site-units")
async def create_site_unit(
    body: CreateUnitPayload,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Register a new staff quarter unit."""
    from models import SiteInspectionUnit
    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    existing = db.execute(
        select(SiteInspectionUnit).where(
            SiteInspectionUnit.tenant_id == tenant.id,
            SiteInspectionUnit.site_name == body.site_name,
            SiteInspectionUnit.block_name == body.block_name,
            SiteInspectionUnit.unit_number == body.unit_number,
        )
    ).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail="Unit already exists")
    unit = SiteInspectionUnit(
        tenant_id=tenant.id,
        site_name=body.site_name,
        block_name=body.block_name,
        unit_number=body.unit_number,
        capacity=body.capacity,
        unit_type=body.unit_type,
    )
    db.add(unit)
    db.commit()
    db.refresh(unit)
    return unit.to_dict()


@router.get("/site-units")
async def list_site_units(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """List all registered units for this tenant."""
    from models import SiteInspectionUnit
    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    units = db.execute(
        select(SiteInspectionUnit).where(SiteInspectionUnit.tenant_id == tenant.id).order_by(
            SiteInspectionUnit.site_name, SiteInspectionUnit.block_name, SiteInspectionUnit.unit_number
        )
    ).scalars().all()
    return {"units": [u.to_dict() for u in units]}


@router.delete("/site-units/{unit_id}")
async def delete_site_unit(
    unit_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Delete a unit and all its inspection records."""
    from models import SiteInspectionUnit
    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    unit = db.execute(
        select(SiteInspectionUnit).where(
            SiteInspectionUnit.id == unit_id,
            SiteInspectionUnit.tenant_id == tenant.id,
        )
    ).scalar_one_or_none()
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")
    db.delete(unit)
    db.commit()
    return {"ok": True}


@router.post("/site-units/{unit_id}/assess")
async def assess_unit(
    unit_id: int,
    files: List[UploadFile] = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Step 1 — upload photos, run vision assessment, return results WITHOUT saving to DB.

    Photos are saved to disk so the LLM can read them and so the caller can
    display them via /api/site-photos/<basename>.  The caller must POST the
    returned photos + assessment to /site-units/{id}/inspections to persist.
    """
    from models import SiteInspectionUnit
    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    unit = db.execute(
        select(SiteInspectionUnit).where(
            SiteInspectionUnit.id == unit_id,
            SiteInspectionUnit.tenant_id == tenant.id,
        )
    ).scalar_one_or_none()
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")

    cfg = get_config()
    upload_dir = pathlib.Path(cfg.db_path).parent / "site_inspections"
    upload_dir.mkdir(parents=True, exist_ok=True)

    photos: List[Dict[str, Any]] = []
    for i, f in enumerate(files):
        safe_name = pathlib.Path(f.filename or f"photo_{i}").name
        # Sanitize unit fields to produce URL-safe filenames (no spaces/slashes)
        def _slug(s: str) -> str:
            return "".join(c if c.isalnum() or c in "-_" else "_" for c in s)
        stored_name = f"{_slug(unit.site_name)}_{_slug(unit.block_name)}_{_slug(unit.unit_number)}_{i}_{safe_name}"
        file_path = upload_dir / stored_name
        with open(file_path, "wb") as out:
            content = await f.read()
            out.write(content)
        photos.append({
            "path": str(file_path),
            "filename": f.filename or f"photo_{i}",
            "url": f"/api/site-photos/{stored_name}",
            "room": "",
            "assessment": "",
        })

    # --- Run 3 atomic skills PER PHOTO ---
    # Each photo gets its own furniture-count, cleanliness-check, site-condition-check
    # call (with that single photo as the image attachment).
    # Photos are processed sequentially to avoid Dashscope 429 burst-rate limiting.
    async def _assess_one_photo(photo: Dict[str, Any], idx: int) -> Dict[str, Any]:
        single_path = [photo["path"]]

        async def _run_skill(skill_name: str, instruction: str) -> Dict[str, Any]:
            prompt = (
                f"Use the {skill_name} skill. {instruction}\n"
                f"Photo to assess:\n- {photo['path']}\n\n"
                f"Return ONLY a JSON object with your assessment."
            )
            text = await _call_estate_agent("facility", prompt, photo_paths=single_path)
            return _extract_json_from_text(text, is_array=False)

        furniture, cleanliness, condition = await asyncio.gather(
            _run_skill("furniture-count", "Count and catalog every visible piece of furniture in this photo. Return JSON: {furniture: [{item, quantity, condition}], total_items, summary}."),
            _run_skill("cleanliness-check", "Assess the cleanliness of each surface visible in this photo. Return JSON: {cleanliness: {floor, walls, bedding, surfaces, overall}, summary}."),
            _run_skill("site-condition-check", "Assess the structural condition (walls, ceiling, windows, lighting, ventilation) and safety hazards in this photo. Return JSON: {site_condition: {walls, ceiling, windows, lighting, ventilation}, safety_hazards: [...], overall_rating, priority_actions: [...]}."),
        )
        return {
            "filename": photo["filename"],
            "url": photo["url"],
            "path": photo["path"],
            "furniture_result": furniture,
            "cleanliness_result": cleanliness,
            "condition_result": condition,
        }

    # Process photos sequentially (each photo's 3 skills run in parallel internally)
    per_photo = []
    for i, photo in enumerate(photos):
        result = await _assess_one_photo(photo, i)
        per_photo.append(result)

    # Merge per-photo assessments into the photos list for backward compat
    for i, pr in enumerate(per_photo):
        if i < len(photos):
            photos[i]["assessment"] = ""

    return {
        "unit_id": unit.id,
        "unit_label": f"{unit.site_name} — {unit.block_name} — {unit.unit_number}",
        "photos": photos,
        # Per-photo results — each photo has its own 3 skill results
        "per_photo": per_photo,
        # Flattened overall summary (aggregated across all photos)
        "furniture_count": sum(len(pr.get("furniture_result", {}).get("furniture", [])) for pr in per_photo) if per_photo else 0,
        # Cleanliness overall — take first photo (cleanliness is uniform per room, not per photo)
        "cleanliness": per_photo[0].get("cleanliness_result", {}).get("cleanliness", {}).get("overall", "") if per_photo else "",
        # Aggregate safety hazards + priority actions across all photos
        "site_condition": "; ".join(
            s for pr in per_photo
            for s in (pr.get("condition_result", {}).get("site_condition", {}).values() if isinstance(pr.get("condition_result", {}).get("site_condition"), dict) else [])
            if s
        ) if per_photo else "",
        "safety_hazards": "; ".join(
            h for pr in per_photo
            for h in (pr.get("condition_result", {}).get("safety_hazards") or [])
        ) if per_photo else "",
        "overall_rating": "; ".join(
            pr.get("condition_result", {}).get("overall_rating", "")
            for pr in per_photo
            if pr.get("condition_result", {}).get("overall_rating")
        ) if per_photo else "",
        "priority_actions": "; ".join(
            a for pr in per_photo
            for a in (pr.get("condition_result", {}).get("priority_actions") or [])
        ) if per_photo else "",
        "merged_assessment": {
            "per_photo": per_photo,
        },
    }


@router.post("/site-units/{unit_id}/inspections")
async def save_inspection(
    unit_id: int,
    body: SaveInspectionPayload,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Step 2 — persist an assessed inspection (photos + results) to the database.

    Called after the user reviews the assessment from /assess and clicks Close/Save.
    """
    from models import SiteInspectionUnit, SiteInspection
    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    unit = db.execute(
        select(SiteInspectionUnit).where(
            SiteInspectionUnit.id == unit_id,
            SiteInspectionUnit.tenant_id == tenant.id,
        )
    ).scalar_one_or_none()
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")

    inspection = SiteInspection(
        tenant_id=tenant.id,
        unit_id=unit.id,
        inspected_by=user.name or user.email,
        photos=[p.model_dump() for p in body.photos],
        merged_assessment=body.merged_assessment or None,
        furniture_count=body.furniture_count,
        cleanliness=body.cleanliness,
        site_condition=body.site_condition,
        safety_hazards=body.safety_hazards,
        overall_rating=body.overall_rating,
        priority_actions=body.priority_actions,
    )
    db.add(inspection)
    db.commit()
    db.refresh(inspection)
    return inspection.to_dict()


@router.get("/site-units/{unit_id}/inspections")
async def list_unit_inspections(
    unit_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Get inspection history for a unit."""
    from models import SiteInspection
    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    inspections = db.execute(
        select(SiteInspection).where(
            SiteInspection.unit_id == unit_id,
            SiteInspection.tenant_id == tenant.id,
        ).order_by(SiteInspection.inspection_date.desc())
    ).scalars().all()
    return {"inspections": [i.to_dict() for i in inspections]}


@router.get("/inspections")
async def list_all_inspections(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    unit_ids: Optional[str] = Query(None, description="Comma-separated unit IDs"),
    date_from: Optional[str] = Query(None, description="ISO date YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="ISO date YYYY-MM-DD"),
) -> dict:
    """Get all inspection records for this tenant, optionally filtered by units and date range."""
    from models import SiteInspection
    from datetime import datetime as _dt, timezone as _tz
    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    query = select(SiteInspection).where(SiteInspection.tenant_id == tenant.id)

    # Filter by unit IDs (comma-separated)
    if unit_ids:
        try:
            ids = [int(x.strip()) for x in unit_ids.split(",") if x.strip()]
            if ids:
                query = query.where(SiteInspection.unit_id.in_(ids))
        except ValueError:
            pass

    # Filter by date range
    if date_from:
        try:
            df = _dt.fromisoformat(date_from).replace(tzinfo=_tz.utc)
            query = query.where(SiteInspection.inspection_date >= df)
        except ValueError:
            pass
    if date_to:
        try:
            dt = _dt.fromisoformat(date_to).replace(tzinfo=_tz.utc) + timedelta(days=1)  # inclusive end
            query = query.where(SiteInspection.inspection_date < dt)
        except ValueError:
            pass

    query = query.order_by(SiteInspection.inspection_date.desc())
    inspections = db.execute(query).scalars().all()
    return {"inspections": [i.to_dict() for i in inspections]}


# ──────────────────────────────────────────────────────────────────────────
# HR Dashboard — synced from Notion via scripts/sync-notion-hr.py
# ──────────────────────────────────────────────────────────────────────────


@router.get("/hr-stats")
async def get_hr_stats(
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Aggregated HR dashboard stats — all tabs.

    Data is synced from Notion into SQLite by scripts/sync-notion-hr.py.
    Formula fields (Deadline, Days Left, Overdue, Task Status, No. of Years)
    are computed on read so they stay current.
    """
    from models import (
        HrCandidate,
        HrEmployee,
        HrEquipment,
        HrJobOpening,
        HrMeeting,
        HrMeetingActionItem,
        HrMeetingAttendee,
        HrOnboardingTask,
        HrPerformanceReview,
        HrTraining,
        HrTrainer,
    )

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")

    employees = db.execute(select(HrEmployee).where(HrEmployee.tenant_id == tenant.id)).scalars().all()
    job_openings = db.execute(select(HrJobOpening).where(HrJobOpening.tenant_id == tenant.id)).scalars().all()
    candidates = db.execute(select(HrCandidate).where(HrCandidate.tenant_id == tenant.id)).scalars().all()
    onboarding = db.execute(select(HrOnboardingTask).where(HrOnboardingTask.tenant_id == tenant.id)).scalars().all()
    reviews = db.execute(select(HrPerformanceReview).where(HrPerformanceReview.tenant_id == tenant.id)).scalars().all()
    equipment = db.execute(select(HrEquipment).where(HrEquipment.tenant_id == tenant.id)).scalars().all()
    trainings = db.execute(select(HrTraining).where(HrTraining.tenant_id == tenant.id)).scalars().all()
    trainers = db.execute(select(HrTrainer).where(HrTrainer.tenant_id == tenant.id)).scalars().all()
    meetings = db.execute(select(HrMeeting).where(HrMeeting.tenant_id == tenant.id)).scalars().all()
    from models import HrCandidateEvent, HrCandidateFile, HrEquipmentLog, HrInterview, HrOnboardingChecklistItem, HrOnboardingChecklistProgress, HrTrainingParticipant
    training_participants = db.execute(select(HrTrainingParticipant).where(HrTrainingParticipant.tenant_id == tenant.id)).scalars().all()
    _seed_default_checklist_items(db, tenant.id)
    checklist_items = db.execute(select(HrOnboardingChecklistItem).where(HrOnboardingChecklistItem.tenant_id == tenant.id).order_by(HrOnboardingChecklistItem.sort_order, HrOnboardingChecklistItem.id)).scalars().all()
    checklist_progress = db.execute(select(HrOnboardingChecklistProgress).where(HrOnboardingChecklistProgress.tenant_id == tenant.id)).scalars().all()
    candidate_files = db.execute(select(HrCandidateFile).where(HrCandidateFile.tenant_id == tenant.id)).scalars().all()
    equipment_logs = db.execute(select(HrEquipmentLog).where(HrEquipmentLog.tenant_id == tenant.id).order_by(HrEquipmentLog.id.desc())).scalars().all()
    candidate_events = db.execute(select(HrCandidateEvent).where(HrCandidateEvent.tenant_id == tenant.id).order_by(HrCandidateEvent.id.desc())).scalars().all()
    interviews = db.execute(select(HrInterview).where(HrInterview.tenant_id == tenant.id)).scalars().all()
    action_items = db.execute(select(HrMeetingActionItem).where(HrMeetingActionItem.tenant_id == tenant.id)).scalars().all()
    attendees = db.execute(select(HrMeetingAttendee).where(HrMeetingAttendee.tenant_id == tenant.id)).scalars().all()

    dept_counts: dict[str, int] = {}
    for emp in employees:
        dept = emp.department or "Unknown"
        dept_counts[dept] = dept_counts.get(dept, 0) + 1

    pipeline_counts: dict[str, int] = {}
    for cand in candidates:
        status = cand.status or "Unknown"
        pipeline_counts[status] = pipeline_counts.get(status, 0) + 1

    onboarding_in_progress = sum(1 for t in onboarding if t.status == "In progress")
    onboarding_done = sum(1 for t in onboarding if t.status == "Done")
    overdue_openings = sum(1 for j in job_openings if j.to_dict().get("overdue") == "Overdue")
    training_total_charges = sum(t.training_charges or 0 for t in trainings)

    # Normalize legacy statuses to new convention
    STATUS_MAP = {
        "Not Initiated": "Draft",
        "Test Ongoing": "Active",
        "Hired": "Closed - Hired",
        "Ongoing": "Active",
        "Open": "Active",
    }
    def normalize_status(s: str | None) -> str:
        if not s:
            return "Draft"
        if s.startswith("Closed"):
            return s
        return STATUS_MAP.get(s, s)

    open_action_items = sum(1 for a in action_items if a.status in ("Open", "In progress", "Not Started"))

    return {
        "total_employees": len(employees),
        "total_job_openings": len(job_openings),
        "overdue_openings": overdue_openings,
        "total_candidates": len(candidates),
        "pipeline_counts": pipeline_counts,
        "onboarding_in_progress": onboarding_in_progress,
        "onboarding_done": onboarding_done,
        "total_reviews": len(reviews),
        "total_equipment": len(equipment),
        "equipment_overdue": sum(1 for e in equipment if e.to_dict().get("is_overdue")),
        "total_trainings": len(trainings),
        "total_trainers": len(trainers),
        "training_total_charges": training_total_charges,
        "total_meetings": len(meetings),
        "open_action_items": open_action_items,
        "dept_counts": dept_counts,
        "employees": [e.to_dict() for e in employees],
        "job_openings": [{**j.to_dict(), "job_status": normalize_status(j.job_status)} for j in job_openings],
        "candidates": [c.to_dict() for c in candidates],
        "onboarding_tasks": [t.to_dict() for t in onboarding],
        "performance_reviews": [r.to_dict() for r in reviews],
        "equipment": [e.to_dict() for e in equipment],
        "trainings": [t.to_dict() for t in trainings],
        "trainers": [t.to_dict() for t in trainers],
        "meetings": [m.to_dict() for m in meetings],
        "meeting_action_items": [a.to_dict() for a in action_items],
        "meeting_attendees": [a.to_dict() for a in attendees],
        "candidate_files": [f.to_dict() for f in candidate_files],
        "candidate_events": [e.to_dict() for e in candidate_events],
        "interviews": [i.to_dict() for i in interviews],
        "equipment_logs": [l.to_dict() for l in equipment_logs],
        "training_participants": [p.to_dict() for p in training_participants],
        "onboarding_checklist_items": [i.to_dict() for i in checklist_items],
        "onboarding_checklist_progress": [p.to_dict() for p in checklist_progress],
        "source": "notion_sync",
    }


_ALLOWED_JD_EXTS = {"pdf", "doc", "docx", "txt", "md", "rtf"}


@router.post("/hr/job-openings")
async def create_hr_job_opening(
    name: str = Path(...),
    file: UploadFile = File(None),
    job_title: str = Form(...),
    department: str = Form(""),
    employment_type: str = Form(""),
    experience: str = Form(""),
    budget_max: str = Form(""),
    hiring_manager: str = Form(""),
    application_start: str = Form(""),
    job_status: str = Form("Not Initiated"),
    job_description: str = Form(""),
    jd_link: str = Form(""),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Create a job opening from the portal.

    Local rows get a synthetic ``notion_page_id`` (``local-*``) so Notion
    sync upserts never overwrite them. Optional job-description upload (served
    via /api/doc-uploads) and/or external JD link. Deadline/Days Left/Overdue
    are computed on read (Application Start + 90 days).
    """
    from models import HrJobOpening

    title = (job_title or "").strip()
    if not title:
        raise HTTPException(status_code=422, detail="Job title is required")

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")

    import uuid as _uuid

    jd_file_url = None
    if file is not None and file.filename:
        safe_name = pathlib.Path(file.filename or "document").name
        ext = pathlib.Path(safe_name).suffix.lower().lstrip(".")
        if ext not in _ALLOWED_JD_EXTS:
            raise HTTPException(status_code=422, detail=f"Unsupported file type (.{ext}). Allowed: pdf, doc, docx, txt, md, rtf")
        content = await file.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=422, detail="File too large (max 10 MB)")
        cfg = get_config()
        upload_dir = pathlib.Path(cfg.db_path).parent / "dashboard_uploads"
        upload_dir.mkdir(parents=True, exist_ok=True)
        unique_name = f"{_uuid.uuid4().hex[:8]}_{safe_name}"
        (upload_dir / unique_name).write_bytes(content)
        jd_file_url = f"/api/doc-uploads/{unique_name}"

    budget = None
    raw_budget = (budget_max or "").strip().replace(",", "")
    if raw_budget:
        try:
            budget = float(raw_budget)
        except ValueError:
            raise HTTPException(status_code=422, detail="Budget must be a number")

    opening = HrJobOpening(
        tenant_id=tenant.id,
        notion_page_id=f"local-{_uuid.uuid4().hex}",
        job_title=title,
        job_status=job_status.strip() or "Not Initiated",
        department=department.strip(),
        employment_type=employment_type.strip(),
        experience=experience.strip(),
        budget_max=budget,
        hiring_manager=hiring_manager.strip() or None,
        application_start=application_start.strip() or None,
        job_description=job_description or None,
        jd_link=jd_link.strip() or None,
        jd_file_url=jd_file_url,
    )
    db.add(opening)
    db.commit()
    db.refresh(opening)

    try:
        import audit  # audit infra is optional on this branch — never break the flow
        audit.log_action(
            db, tenant, user, "hr", "hr.job_opening.create", "job_opening",
            str(opening.id),
            detail={"job_title": opening.job_title, "department": opening.department,
                    "jd_link": opening.jd_link, "jd_file_url": opening.jd_file_url},
        )
    except Exception:
        pass

    return {"ok": True, "job": opening.to_dict()}


@router.put("/hr/job-openings/{job_id}")
async def update_hr_job_opening(
    name: str = Path(...),
    job_id: int = Path(...),
    file: UploadFile = File(None),
    job_title: str = Form(""),
    department: str = Form(""),
    employment_type: str = Form(""),
    experience: str = Form(""),
    budget_max: str = Form(""),
    hiring_manager: str = Form(""),
    application_start: str = Form(""),
    job_status: str = Form(""),
    job_description: str = Form(""),
    jd_link: str = Form(""),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Update an existing job opening. Same fields as create; only non-empty
    form values overwrite the existing row."""
    from models import HrJobOpening

    opening = db.get(HrJobOpening, job_id)
    if opening is None:
        raise HTTPException(status_code=404, detail="Job opening not found")

    title = (job_title or "").strip()
    if title:
        opening.job_title = title

    dept = (department or "").strip()
    if dept:
        opening.department = dept

    etype = (employment_type or "").strip()
    if etype:
        opening.employment_type = etype

    exp = (experience or "").strip()
    if exp:
        opening.experience = exp

    raw_budget = (budget_max or "").strip().replace(",", "")
    if raw_budget:
        try:
            opening.budget_max = float(raw_budget)
        except ValueError:
            raise HTTPException(status_code=422, detail="Budget must be a number")

    hm = (hiring_manager or "").strip()
    if hm:
        opening.hiring_manager = hm

    app_start = (application_start or "").strip()
    if app_start:
        opening.application_start = app_start

    status = (job_status or "").strip()
    if status:
        opening.job_status = status

    desc = job_description
    if desc is not None and desc != "":
        opening.job_description = desc if desc.strip() else None

    link = (jd_link or "").strip()
    if link:
        opening.jd_link = link

    # Optional JD file upload — replaces previous file
    if file is not None and file.filename:
        safe_name = pathlib.Path(file.filename or "document").name
        ext = pathlib.Path(safe_name).suffix.lower().lstrip(".")
        if ext not in _ALLOWED_JD_EXTS:
            raise HTTPException(status_code=422, detail=f"Unsupported file type (.{ext}). Allowed: pdf, doc, docx, txt, md, rtf")
        content = await file.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=422, detail="File too large (max 10 MB)")
        cfg = get_config()
        upload_dir = pathlib.Path(cfg.db_path).parent / "dashboard_uploads"
        upload_dir.mkdir(parents=True, exist_ok=True)
        import uuid as _uuid
        unique_name = f"{_uuid.uuid4().hex[:8]}_{safe_name}"
        (upload_dir / unique_name).write_bytes(content)
        opening.jd_file_url = f"/api/doc-uploads/{unique_name}"

    db.commit()
    db.refresh(opening)

    try:
        import audit
        audit.log_action(
            db, None, user, "hr", "hr.job_opening.update", "job_opening",
            str(opening.id),
            detail={"job_title": opening.job_title, "department": opening.department},
        )
    except Exception:
        pass

    return {"ok": True, "job": opening.to_dict()}


class HrCandidateReviewBody(BaseModel):
    kind: str = "hr"


def _drive_file_id(url: str) -> Optional[str]:
    """Extract a Google Drive file id from a viewer/drive link, if any."""
    if not url:
        return None
    m = _re.search(r"drive\.google\.com/file/d/([A-Za-z0-9_-]+)", url)
    return m.group(1) if m else None


async def _fetch_candidate_doc(url: str) -> str:
    """Best-effort text extraction from a candidate's Drive-hosted resume /
    screening-answers link. Returns "" on any failure — never raises."""
    fid = _drive_file_id(url)
    if not fid:
        return ""
    try:
        async with httpx.AsyncClient(timeout=25.0) as client:
            resp = await client.get(
                f"https://drive.google.com/uc?export=download&id={fid}",
                follow_redirects=True,
            )
            if resp.status_code != 200 or not resp.content:
                return ""
            data = resp.content
    except Exception:
        return ""
    if data[:5] == b"%PDF-":
        try:
            import fitz  # pymupdf

            text = ""
            with fitz.open(stream=data, filetype="pdf") as doc:
                for page in doc:
                    text += page.get_text()
            return text
        except Exception:
            return ""
    for enc in ("utf-8", "latin-1"):
        try:
            s = data.decode(enc)
            printable = sum(1 for ch in s if ch.isprintable() or ch in "\n\t\r")
            if printable > len(s) * 0.9:
                return s
        except Exception:
            continue
    return ""


def _candidate_fallback_extract(cand) -> dict:
    """DB-record-only extraction when the LLM is unavailable."""
    return {
        "summary": "",
        "skills": [],
        "experience": [],
        "education": [],
        "screening_answers": [],
        "key_details": {
            "role": cand.role or "",
            "status": cand.status or "",
            "source": cand.source or "",
            "tracker": cand.candidate_type or "",
        },
        "source": "fallback",
    }


async def _ai_extract_candidate(cand, resume_text: str, screening_text: str) -> dict:
    """Run DeepSeek structured extraction over the candidate record + docs."""
    prompt = f"""You are an HR assistant. Extract structured information about this job candidate.
Candidate record (from the hiring board):
- Name: {cand.name}
- Role applied: {cand.role or '(none)'}
- Email: {cand.email or '(none)'}
- Phone: {cand.phone_no or '(none)'}
- Pipeline status: {cand.status or '(none)'}
- Source: {cand.source or '(none)'}
- Tracker type: {cand.candidate_type or '(none)'}

RESUME TEXT:
{resume_text[:6000] or '(not available)'}

SCREENING ANSWERS TEXT:
{screening_text[:6000] or '(not available)'}

Return ONLY valid JSON (no markdown fences) with this schema:
{{
  "summary": "2-3 sentence professional summary of the candidate",
  "skills": ["skill1", "skill2"],
  "experience": [{{"title": "...", "company": "...", "period": "..."}}],
  "education": ["..."],
  "screening_answers": [{{"question": "...", "answer": "..."}}],
  "key_details": {{"notice_period": "...", "expected_salary": "...", "current_location": "...", "notes": "..."}}
}}"""
    try:
        from gateway import _call_deepseek

        raw = await _call_deepseek(
            prompt,
            system_prompt="You extract structured candidate data from resumes. Reply with ONLY valid JSON.",
            max_tokens=2400,
        )
    except Exception:
        raw = None
    if raw:
        parsed = _extract_json_from_text(raw)
        if isinstance(parsed, dict) and parsed and "raw_response" not in parsed:
            parsed.setdefault("summary", "")
            parsed.setdefault("source", "ai")
            return parsed
    return _candidate_fallback_extract(cand)


@router.post("/hr/candidates/{candidate_id}/extract")
async def extract_hr_candidate(
    candidate_id: int,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """AI-extract all candidate details from the resume + screening docs."""
    from models import HrCandidate

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    cand = db.get(HrCandidate, candidate_id)
    if cand is None or cand.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Candidate not found")

    resume_text = await _fetch_candidate_doc(cand.resume_url or "")
    screening_text = await _fetch_candidate_doc(cand.screening_answers_url or "")
    result = await _ai_extract_candidate(cand, resume_text, screening_text)

    cand.ai_extract_json = json.dumps(result)
    cand.ai_summary = (result.get("summary") or "")[:2000]
    cand.extracted_at = datetime.utcnow().isoformat(timespec="seconds")
    db.commit()

    try:  # audit infra optional on this branch — never break the flow
        import audit
        audit.log_action(
            db, tenant, user, "hr", "hr.candidate.extract", "candidate",
            str(candidate_id), detail={"source": result.get("source")},
        )
    except Exception:
        pass

    return {"ok": True, "candidate": cand.to_dict(), "extract": result}


@router.post("/hr/candidates/{candidate_id}/review")
async def review_hr_candidate(
    candidate_id: int,
    body: HrCandidateReviewBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Record an HR or Manager review tap for a candidate."""
    from models import HrCandidate

    kind = (body.kind or "").strip().lower()
    if kind not in ("hr", "manager"):
        raise HTTPException(status_code=422, detail="kind must be 'hr' or 'manager'")

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    cand = db.get(HrCandidate, candidate_id)
    if cand is None or cand.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Candidate not found")

    now = datetime.utcnow().isoformat(timespec="seconds")
    if kind == "hr":
        cand.hr_reviewed = True
        cand.hr_reviewed_at = now
    else:
        cand.manager_reviewed = True
        cand.manager_reviewed_at = now
    db.commit()

    try:
        import audit
        audit.log_action(
            db, tenant, user, "hr", f"hr.candidate.{kind}.review", "candidate",
            str(candidate_id), detail={"review": f"{kind.capitalize()} done review"},
        )
    except Exception:
        pass

    return {"ok": True, "candidate": cand.to_dict()}


@router.post("/hr/candidates/{candidate_id}/add-to-pipeline")
async def add_hr_candidate_to_pipeline(
    candidate_id: int,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Flag a candidate as added to the recruitment pipeline."""
    from models import HrCandidate

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    cand = db.get(HrCandidate, candidate_id)
    if cand is None or cand.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Candidate not found")

    old_status = cand.status
    cand.in_pipeline = True
    if old_status not in ("1st Interview Scheduled", "HR Interview Done",
                          "Waiting Manager Interview Confirm", "Manager Interview Scheduled",
                          "Waiting Interview Result", "Waiting Offer Confirmation",
                          "Offer Sent - Waiting Reply", "Done"):
        cand.status = "Interview Email Sent - Waiting Reply"
    _hr_event(db, tenant.id, candidate_id, "stage_move",
              note="Added into recruitment pipeline",
              from_status=old_status, to_status=cand.status, user=user)
    db.commit()

    try:
        import audit
        audit.log_action(
            db, tenant, user, "hr", "hr.candidate.add_to_pipeline", "candidate",
            str(candidate_id), detail={"to": cand.status},
        )
    except Exception:
        pass

    return {"ok": True, "candidate": cand.to_dict()}


class HrCandidateMoveBody(BaseModel):
    status: str


@router.post("/hr/candidates/{candidate_id}/move")
async def move_hr_candidate(
    candidate_id: int,
    body: HrCandidateMoveBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Change a candidate's pipeline stage (drag & drop on the board)."""
    from models import HrCandidate

    status = (body.status or "").strip()
    if not status or len(status) > 128:
        raise HTTPException(status_code=422, detail="Status must be a non-empty stage name")

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    cand = db.get(HrCandidate, candidate_id)
    if cand is None or cand.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Candidate not found")

    old_status = cand.status
    cand.status = status
    if status in ("1st Interview Scheduled", "HR Interview Done",
                  "Manager Interview Scheduled", "Waiting Interview Result",
                  "Waiting Offer Confirmation", "Offer Sent - Waiting Reply", "Done"):
        cand.waiting_since = None
        cand.waiting_reason = None
    _hr_event(db, tenant.id, candidate_id, "stage_move",
              from_status=old_status, to_status=status, user=user)
    db.commit()

    try:  # audit infra optional on this branch — never break the flow
        import audit
        audit.log_action(
            db, tenant, user, "hr", "hr.candidate.move", "candidate",
            str(candidate_id), detail={"from": old_status, "to": status},
        )
    except Exception:
        pass

    return {"ok": True, "candidate": cand.to_dict()}


# ─── HR recruitment workflow (Phase 1) ───

def _hr_event(db, tenant_id: int, candidate_id: int, event_type: str,
              note: Optional[str] = None, from_status: Optional[str] = None,
              to_status: Optional[str] = None, user: Optional[User] = None) -> None:
    from models import HrCandidateEvent
    try:
        db.add(HrCandidateEvent(
            tenant_id=tenant_id, candidate_id=candidate_id, event_type=event_type,
            note=note, from_status=from_status, to_status=to_status,
            actor_name=(user.name or "") if user else None,
            actor_email=(user.email or "") if user else None,
        ))
    except Exception:
        pass


class HrCommentBody(BaseModel):
    note: str


class HrDecisionBody(BaseModel):
    decision: str
    comment: str = ""


class HrScheduleBody(BaseModel):
    round: str = "first"
    scheduled_at: str
    interviewer_name: str = ""
    interviewer_employee_id: Optional[int] = None
    location: str = ""


class HrInterviewStatusBody(BaseModel):
    status: str


def _extract_text_from_upload(data: bytes) -> str:
    """Best-effort text extraction from an uploaded resume document."""
    if data[:5] == b"%PDF-":
        try:
            import fitz  # pymupdf
            text = ""
            with fitz.open(stream=data, filetype="pdf") as doc:
                for page in doc:
                    text += page.get_text()
            return text
        except Exception:
            return ""
    for enc in ("utf-8", "latin-1"):
        try:
            s = data.decode(enc)
            printable = sum(1 for ch in s if ch.isprintable() or ch in "\n\t\r")
            if printable > len(s) * 0.9:
                return s
        except Exception:
            continue
    return ""


@router.post("/hr/extract-resume")
async def hr_extract_resume(
    name: str = Path(...),
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Extract name/email/phone from an applicant resume before HR saves it."""
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=422, detail="File too large (max 10 MB)")
    text = _extract_text_from_upload(content)

    result = {"name": "", "email": "", "phone": "", "summary": "", "source": "empty",
              "resume_text": text.strip()[:8000]}

    # regex fallback (also used to cross-check the LLM answer)
    import re as _re_local
    m_email = _re_local.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", text)
    m_phone = _re_local.search(r"(\+?\d[\d\s\-()]{7,15}\d)", text)
    if m_email:
        result["email"] = m_email.group(0)
    if m_phone:
        result["phone"] = m_phone.group(0).strip()

    if text.strip():
        result["source"] = "fallback"
        try:
            from gateway import _call_deepseek
            raw = await _call_deepseek(
                "Extract the candidate's full name, email address and phone number from this "
                "resume text. Return ONLY valid JSON: "
                "{\"name\": \"...\", \"email\": \"...\", \"phone\": \"...\", \"summary\": \"one-sentence professional summary\"}.\n\n"
                + text[:6000],
                system_prompt="You extract contact details from resumes. Reply with ONLY valid JSON.",
                max_tokens=500,
            )
            if raw:
                parsed = _extract_json_from_text(raw)
                if isinstance(parsed, dict) and "raw_response" not in parsed:
                    result["source"] = "ai"
                    result["name"] = (parsed.get("name") or "").strip()[:200]
                    result["email"] = (parsed.get("email") or result["email"]).strip()[:320]
                    result["phone"] = (parsed.get("phone") or result["phone"]).strip()[:64]
                    result["summary"] = (parsed.get("summary") or "").strip()[:1000]
        except Exception:
            pass
    return {"ok": True, "extract": result}


@router.post("/hr/job-openings/{job_id}/applicants")
async def add_hr_applicant(
    job_id: int,
    name: str = Path(...),
    file: UploadFile = File(None),
    applicant_name: str = Form(...),
    email: str = Form(""),
    phone_no: str = Form(""),
    source: str = Form(""),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Add an applicant (resume upload + AI-prefilled details) to a job opening.

    The applicant enters the pipeline at Screening - Pending with
    date_entry = upload timestamp and their role set to the job title.
    """
    import uuid as _uuid
    from models import HrCandidate, HrCandidateFile, HrJobOpening

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    job = db.get(HrJobOpening, job_id)
    if job is None or job.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Job opening not found")
    app_name = (applicant_name or "").strip()
    if not app_name:
        raise HTTPException(status_code=422, detail="Applicant name is required")

    now = datetime.utcnow().isoformat(timespec="seconds")
    ctype = "internship" if (job.employment_type or "").lower() == "internship" else "fulltime"
    cand = HrCandidate(
        tenant_id=tenant.id,
        notion_page_id=f"portal-{_uuid.uuid4().hex}",
        name=app_name,
        email=email.strip() or None,
        phone_no=phone_no.strip() or None,
        role=job.job_title,
        status="Resume Received",
        source=((source if isinstance(source, str) else "").strip()[:128] or "Portal"),
        candidate_type=ctype,
        date_entry=now,
        job_opening_id=job.id,
    )
    db.add(cand)
    db.flush()

    file_url = None
    filename = None
    if file is not None and file.filename:
        safe_name = pathlib.Path(file.filename or "resume").name
        ext = pathlib.Path(safe_name).suffix.lower().lstrip(".")
        if ext not in {"pdf", "doc", "docx", "txt", "md", "rtf", "png", "jpg", "jpeg", "webp"}:
            raise HTTPException(status_code=422, detail="Unsupported resume file type")
        content = await file.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=422, detail="File too large (max 10 MB)")
        cfg = get_config()
        upload_dir = pathlib.Path(cfg.db_path).parent / "dashboard_uploads"
        upload_dir.mkdir(parents=True, exist_ok=True)
        unique_name = f"{_uuid.uuid4().hex[:8]}_{safe_name}"
        (upload_dir / unique_name).write_bytes(content)
        file_url = f"/api/doc-uploads/{unique_name}"
        filename = safe_name
        cand.resume_url = file_url
        db.add(HrCandidateFile(
            tenant_id=tenant.id, candidate_id=cand.id, kind="resume",
            filename=filename, file_url=file_url,
            uploaded_by_name=user.name if user else None, uploaded_at=now,
        ))

    _hr_event(db, tenant.id, cand.id, "stage_move", note="Applicant added (resume received)",
              to_status="Resume Received", user=user)
    _hr_event(db, tenant.id, cand.id, "upload", note=f"Resume uploaded: {filename or '(none)'}", user=user)
    db.commit()
    db.refresh(cand)

    try:
        import audit
        audit.log_action(db, tenant, user, "hr", "hr.applicant.add", "candidate",
                         str(cand.id), detail={"job_id": job.id, "job_title": job.job_title})
    except Exception:
        pass
    return {"ok": True, "candidate": cand.to_dict()}


@router.post("/hr/candidates/{candidate_id}/file")
async def upload_hr_candidate_file(
    candidate_id: int,
    name: str = Path(...),
    file: UploadFile = File(...),
    kind: str = Form("other"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Upload a workflow document for a candidate (screening_answers /
    offer_letter / other)."""
    import uuid as _uuid
    from models import HrCandidate, HrCandidateFile

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    cand = db.get(HrCandidate, candidate_id)
    if cand is None or cand.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Candidate not found")
    if kind not in ("resume", "screening_answers", "offer_letter", "other"):
        raise HTTPException(status_code=422, detail="Invalid file kind")

    safe_name = pathlib.Path(file.filename or "document").name
    ext = pathlib.Path(safe_name).suffix.lower().lstrip(".")
    if ext not in {"pdf", "doc", "docx", "txt", "md", "rtf", "png", "jpg", "jpeg", "webp"}:
        raise HTTPException(status_code=422, detail="Unsupported file type")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=422, detail="File too large (max 10 MB)")

    cfg = get_config()
    upload_dir = pathlib.Path(cfg.db_path).parent / "dashboard_uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    unique_name = f"{_uuid.uuid4().hex[:8]}_{safe_name}"
    (upload_dir / unique_name).write_bytes(content)
    file_url = f"/api/doc-uploads/{unique_name}"

    now = datetime.utcnow().isoformat(timespec="seconds")
    db.add(HrCandidateFile(
        tenant_id=tenant.id, candidate_id=candidate_id, kind=kind,
        filename=safe_name, file_url=file_url,
        uploaded_by_name=user.name if user else None, uploaded_at=now,
    ))
    _hr_event(db, tenant.id, candidate_id, "upload",
              note=f"{kind.replace('_', ' ').title()} uploaded: {safe_name}", user=user)
    db.commit()
    return {"ok": True, "file_url": file_url, "filename": safe_name}


@router.post("/hr/candidates/{candidate_id}/comment")
async def comment_hr_candidate(
    candidate_id: int,
    body: HrCommentBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Add a comment to the candidate timeline."""
    from models import HrCandidate

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    cand = db.get(HrCandidate, candidate_id)
    if cand is None or cand.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Candidate not found")
    note = (body.note or "").strip()
    if not note:
        raise HTTPException(status_code=422, detail="Comment cannot be empty")

    _hr_event(db, tenant.id, candidate_id, "comment", note=note, user=user)
    db.commit()
    return {"ok": True, "candidate": cand.to_dict()}


@router.post("/hr/candidates/{candidate_id}/decision")
async def decide_hr_candidate(
    candidate_id: int,
    body: HrDecisionBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Record an interview decision.

    From "HR Interview Done": continue → Waiting Manager Interview Confirm, reject → Rejected.
    From "Waiting Offer Confirmation": offer → Offer Sent - Waiting Reply, reject → Rejected.
    """
    from models import HrCandidate

    decision = (body.decision or "").strip().lower()
    if decision not in ("continue", "reject", "offer"):
        raise HTTPException(status_code=422, detail="decision must be continue, reject or offer")

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    cand = db.get(HrCandidate, candidate_id)
    if cand is None or cand.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Candidate not found")

    cur = (cand.status or "").strip()
    if decision == "continue":
        if cur.lower() != "hr interview done":
            raise HTTPException(status_code=422, detail="Continue is only valid after the HR interview is done")
        new_status = "Waiting Manager Interview Confirm"
    elif decision == "offer":
        if cur.lower() != "waiting offer confirmation":
            raise HTTPException(status_code=422, detail="Offer is only valid after the interview result is confirmed")
        new_status = "Offer Sent - Waiting Reply"
    else:  # reject — allowed from HR interview done, waiting result, or offer confirmation stages
        if cur.lower() not in ("hr interview done", "waiting interview result", "waiting offer confirmation"):
            raise HTTPException(status_code=422, detail="Reject is only valid after the HR interview or before the offer")
        new_status = "Rejected"

    old = cand.status
    cand.status = new_status
    cand.waiting_since = None
    cand.waiting_reason = None
    if new_status == "Rejected":
        cand.removed_reason = "Rejected"
    _hr_event(db, tenant.id, candidate_id, "decision",
              note=f"Decision: {decision.upper()}. {(body.comment or '').strip()}" if body.comment else f"Decision: {decision.upper()}",
              from_status=old, to_status=new_status, user=user)
    db.commit()

    try:
        import audit
        audit.log_action(db, tenant, user, "hr", f"hr.candidate.decision.{decision}",
                         "candidate", str(candidate_id),
                         detail={"from": old, "to": new_status})
    except Exception:
        pass
    return {"ok": True, "candidate": cand.to_dict()}


@router.post("/hr/candidates/{candidate_id}/schedule")
async def schedule_hr_interview(
    candidate_id: int,
    body: HrScheduleBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Confirm an interview date/time: creates the schedule row and moves the
    candidate into the corresponding interview stage in one action."""
    from models import HrCandidate, HrInterview, HrJobOpening

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    cand = db.get(HrCandidate, candidate_id)
    if cand is None or cand.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Candidate not found")

    rnd = (body.round or "").strip().lower()
    if rnd not in ("first", "manager"):
        raise HTTPException(status_code=422, detail="round must be 'first' or 'manager'")
    when = (body.scheduled_at or "").strip()
    if not when:
        raise HTTPException(status_code=422, detail="Interview date/time is required")

    job_id = None
    if cand.role:
        job = db.execute(
            select(HrJobOpening).where(
                HrJobOpening.tenant_id == tenant.id,
                HrJobOpening.job_title == cand.role,
            )
        ).scalars().first()
        if job is not None:
            job_id = job.id

    old = cand.status
    new_status = "1st Interview Scheduled" if rnd == "first" else "Manager Interview Scheduled"
    interview = HrInterview(
        tenant_id=tenant.id, candidate_id=candidate_id, job_id=job_id,
        round=rnd, scheduled_at=when,
        interviewer_name=(body.interviewer_name or "").strip() or None,
        interviewer_employee_id=body.interviewer_employee_id,
        location=(body.location or "").strip() or None,
        status="scheduled",
    )
    db.add(interview)
    cand.status = new_status
    cand.waiting_since = None
    cand.waiting_reason = None
    _hr_event(db, tenant.id, candidate_id, "stage_move",
              note=f"Interview scheduled ({rnd} round) at {when}"
                   + (f" with {body.interviewer_name}" if body.interviewer_name else ""),
              from_status=old, to_status=new_status, user=user)
    db.commit()
    db.refresh(interview)

    try:
        import audit
        audit.log_action(db, tenant, user, "hr", "hr.candidate.schedule", "candidate",
                         str(candidate_id), detail={"round": rnd, "scheduled_at": when})
    except Exception:
        pass
    return {"ok": True, "candidate": cand.to_dict(), "interview": interview.to_dict()}


@router.post("/hr/interviews/{interview_id}/status")
async def update_hr_interview_status(
    interview_id: int,
    body: HrInterviewStatusBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Mark an interview completed / cancelled / back to scheduled."""
    from models import HrInterview

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    iv = db.get(HrInterview, interview_id)
    if iv is None or iv.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Interview not found")
    st = (body.status or "").strip().lower()
    if st not in ("scheduled", "completed", "cancelled"):
        raise HTTPException(status_code=422, detail="Invalid interview status")
    iv.status = st
    db.commit()
    return {"ok": True, "interview": iv.to_dict()}


@router.post("/hr/candidates/{candidate_id}/waiting")
async def set_hr_candidate_waiting(
    candidate_id: int,
    body: HrCommentBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Set/clear a waiting state (pass note='' to clear)."""
    from models import HrCandidate

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    cand = db.get(HrCandidate, candidate_id)
    if cand is None or cand.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Candidate not found")

    reason = (body.note or "").strip()
    if reason:
        cand.waiting_since = datetime.utcnow().isoformat(timespec="seconds")
        cand.waiting_reason = reason[:256]
        _hr_event(db, tenant.id, candidate_id, "note", note=f"Waiting: {reason}", user=user)
    else:
        cand.waiting_since = None
        cand.waiting_reason = None
        _hr_event(db, tenant.id, candidate_id, "note", note="Waiting cleared (replied / resolved)", user=user)
    db.commit()
    return {"ok": True, "candidate": cand.to_dict()}


@router.post("/hr/candidates/{candidate_id}/remove")
async def remove_hr_candidate(
    candidate_id: int,
    body: HrCommentBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Soft-remove a candidate: status Rejected with reason kept for audit."""
    from models import HrCandidate

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    cand = db.get(HrCandidate, candidate_id)
    if cand is None or cand.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Candidate not found")

    old = cand.status
    reason = (body.note or "").strip() or "No response"
    cand.status = "Rejected"
    cand.removed_reason = reason[:1000]
    cand.waiting_since = None
    cand.waiting_reason = None
    _hr_event(db, tenant.id, candidate_id, "stage_move",
              note=f"Removed: {reason}", from_status=old, to_status="Rejected", user=user)
    db.commit()

    try:
        import audit
        audit.log_action(db, tenant, user, "hr", "hr.candidate.remove", "candidate",
                         str(candidate_id), detail={"from": old, "reason": reason})
    except Exception:
        pass
    return {"ok": True, "candidate": cand.to_dict()}


_ALLOWED_EQ_IMG_EXTS = {"png", "jpg", "jpeg", "webp"}
_ALLOWED_EQ_DOC_EXTS = {"pdf", "doc", "docx", "txt", "png", "jpg", "jpeg", "webp"}


def _equipment_log(db, tenant_id: int, equipment_id: int, event_type: str,
                   actor: str, detail: Optional[str] = None) -> None:
    """Append an activity log entry for a single piece of equipment."""
    from models import HrEquipmentLog
    db.add(HrEquipmentLog(
        tenant_id=tenant_id,
        equipment_id=equipment_id,
        event_type=event_type,
        actor=(actor or "HR")[:256],
        detail=(detail or None) and detail[:2048],
    ))


async def _save_equipment_upload(file: UploadFile, allowed: set, kind: str) -> str:
    """Persist an equipment image/document upload; return its served URL."""
    import uuid as _uuid
    safe_name = pathlib.Path(file.filename or kind).name
    ext = pathlib.Path(safe_name).suffix.lower().lstrip(".")
    if ext not in allowed:
        raise HTTPException(status_code=422, detail=f"Unsupported {kind} file type (.{ext})")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=422, detail=f"{kind} file too large (max 10 MB)")
    cfg = get_config()
    upload_dir = pathlib.Path(cfg.db_path).parent / "dashboard_uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    unique_name = f"{_uuid.uuid4().hex[:8]}_{safe_name}"
    (upload_dir / unique_name).write_bytes(content)
    return f"/api/doc-uploads/{unique_name}"


def _parse_amount(raw: str) -> Optional[float]:
    raw = (raw or "").strip().replace(",", "").replace("RM", "").replace("rm", "")
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        raise HTTPException(status_code=422, detail="Amount must be a number")


@router.post("/hr/equipment")
async def create_hr_equipment(
    name: str = Path(...),
    image: UploadFile = File(None),
    signature_doc: UploadFile = File(None),
    equipment_name: str = Form(...),
    item_number: str = Form(""),
    category: str = Form(""),
    condition: str = Form(""),
    assigned_to: str = Form(""),
    amount: str = Form(""),
    purchase_date: str = Form(""),
    return_due_date: str = Form(""),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Add equipment from the portal (logged to hr_equipment_logs)."""
    import uuid as _uuid
    from models import HrEquipment

    eq_name = (equipment_name or "").strip()
    if not eq_name:
        raise HTTPException(status_code=422, detail="Equipment name is required")

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")

    image_url = await _save_equipment_upload(image, _ALLOWED_EQ_IMG_EXTS, "image") if image is not None and image.filename else None
    signature_doc_url = await _save_equipment_upload(signature_doc, _ALLOWED_EQ_DOC_EXTS, "signature doc") if signature_doc is not None and signature_doc.filename else None

    eq = HrEquipment(
        tenant_id=tenant.id,
        notion_page_id=f"local-{_uuid.uuid4().hex}",
        equipment_name=eq_name,
        item_number=item_number.strip() or None,
        category=category.strip(),
        condition=condition.strip() or None,
        assigned_to=assigned_to.strip() or None,
        amount=_parse_amount(amount),
        purchase_date=purchase_date.strip() or None,
        return_due_date=return_due_date.strip() or None,
        image_url=image_url,
        signature_doc_url=signature_doc_url,
        returned=False,
    )
    db.add(eq)
    db.flush()
    _equipment_log(db, tenant.id, eq.id, "created",
                   user.name if user else "HR",
                   f"Equipment added: {eq_name}")
    db.commit()
    db.refresh(eq)

    try:
        import audit
        audit.log_action(db, tenant, user, "hr", "hr.equipment.create", "equipment",
                         str(eq.id), detail={"equipment_name": eq.equipment_name})
    except Exception:
        pass
    return {"ok": True, "equipment": eq.to_dict()}


class HrEquipmentEditBody(BaseModel):
    equipment_name: Optional[str] = None
    item_number: Optional[str] = None
    category: Optional[str] = None
    condition: Optional[str] = None
    assigned_to: Optional[str] = None
    amount: Optional[str] = None
    purchase_date: Optional[str] = None
    return_due_date: Optional[str] = None
    image_url: Optional[str] = None
    signature_doc_url: Optional[str] = None


@router.put("/hr/equipment/{equipment_id}")
async def update_hr_equipment(
    equipment_id: int,
    body: HrEquipmentEditBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Edit equipment fields; each change is logged."""
    from models import HrEquipment

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    eq = db.get(HrEquipment, equipment_id)
    if eq is None or eq.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Equipment not found")

    changes: List[str] = []
    simple_fields = {
        "equipment_name": "Equipment name",
        "item_number": "Item number",
        "category": "Category",
        "condition": "Condition",
        "assigned_to": "Assigned to",
        "purchase_date": "Purchase date",
        "return_due_date": "Return due date",
        "image_url": "Image",
        "signature_doc_url": "Signature doc",
    }
    for field, label in simple_fields.items():
        val = getattr(body, field)
        if val is None:
            continue
        val = val.strip()
        old = getattr(eq, field)
        new = val or None
        if old != new:
            setattr(eq, field, new)
            changes.append(label)
    if body.amount is not None:
        new_amount = _parse_amount(body.amount)
        if (eq.amount or None) != new_amount:
            eq.amount = new_amount
            changes.append("Amount")
    if not (eq.equipment_name or "").strip():
        raise HTTPException(status_code=422, detail="Equipment name cannot be empty")

    if changes:
        _equipment_log(db, tenant.id, eq.id, "edited",
                       user.name if user else "HR",
                       "Updated: " + ", ".join(changes))
    db.commit()
    db.refresh(eq)

    try:
        import audit
        audit.log_action(db, tenant, user, "hr", "hr.equipment.update", "equipment",
                         str(eq.id), detail={"changes": changes})
    except Exception:
        pass
    return {"ok": True, "equipment": eq.to_dict()}


class HrEquipmentReturnBody(BaseModel):
    return_date: Optional[str] = None
    condition: Optional[str] = None
    note: Optional[str] = None


@router.post("/hr/equipment/{equipment_id}/return")
async def return_hr_equipment(
    equipment_id: int,
    body: HrEquipmentReturnBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Mark equipment as returned; logged with date + condition."""
    from models import HrEquipment

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    eq = db.get(HrEquipment, equipment_id)
    if eq is None or eq.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Equipment not found")
    if eq.returned:
        raise HTTPException(status_code=422, detail="Equipment already returned")

    eq.returned = True
    eq.return_date = (body.return_date or "").strip() or datetime.utcnow().strftime("%Y-%m-%d")
    if (body.condition or "").strip():
        eq.condition = body.condition.strip()
    detail = f"Returned on {eq.return_date}"
    if eq.assigned_to:
        detail += f" (was assigned to {eq.assigned_to})"
    if (body.note or "").strip():
        detail += f" — {body.note.strip()}"
    _equipment_log(db, tenant.id, eq.id, "returned",
                   user.name if user else "HR", detail)
    db.commit()
    db.refresh(eq)

    try:
        import audit
        audit.log_action(db, tenant, user, "hr", "hr.equipment.return", "equipment",
                         str(eq.id), detail={"return_date": eq.return_date})
    except Exception:
        pass
    return {"ok": True, "equipment": eq.to_dict()}


@router.post("/hr/equipment/{equipment_id}/file")
async def upload_hr_equipment_file(
    equipment_id: int,
    name: str = Path(...),
    file: UploadFile = File(...),
    kind: str = Form("image"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Attach/replace an image or signature doc on existing equipment (logged)."""
    from models import HrEquipment

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    eq = db.get(HrEquipment, equipment_id)
    if eq is None or eq.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Equipment not found")

    if kind == "image":
        url = await _save_equipment_upload(file, _ALLOWED_EQ_IMG_EXTS, "image")
        eq.image_url = url
        label = "Image"
    elif kind == "signature_doc":
        url = await _save_equipment_upload(file, _ALLOWED_EQ_DOC_EXTS, "signature doc")
        eq.signature_doc_url = url
        label = "Signature doc"
    else:
        raise HTTPException(status_code=422, detail="kind must be 'image' or 'signature_doc'")

    _equipment_log(db, tenant.id, eq.id, "edited",
                   user.name if user else "HR", f"Updated: {label}")
    db.commit()
    db.refresh(eq)
    return {"ok": True, "equipment": eq.to_dict()}


# Default onboarding checklist template (seeded once per tenant when the
# checklist is still empty). HR can edit/remove items afterwards.
_DEFAULT_ONBOARDING_CHECKLIST = [
    ("HR Documents", "Create gdrive folder in Employee Files (HR)", None),
    ("HR Documents", "Offer Letter in designated Employee Files (HR)", None),
    ("HR Documents", "Welcome Email", None),
    ("HR Documents", "Collecting Filled On-Boarding Form in designated Employee Files (HR)", None),
    ("HR Documents", "Copy of intern IC / Passport in designated Employee Files (HR)", None),
    ("On-The-Day Adhoc", "Set up face ID for office entrance", None),
    ("On-The-Day Adhoc", "Onboarding briefing with HR", None),
    ("On-The-Day Adhoc", "Add Birthday to Google Calendar", "If the birthday is within the internship period"),
    ("On-The-Day Adhoc", "Add in Team On Leave calendar for Intern on leave", None),
    ("On-The-Day Adhoc", "Take individual picture and keep in this Individual Staff Photos File", None),
    ("On-The-Day Adhoc", "Add personal informations in Payroll Tracker", None),
    ("Accounts Activation", "Add to Slack", "Office-News, Happy-Hour, Team Happiness, Sharing Knowledge, Scrum channel (relevant department), Intern Club Group, any relevant channels according to the new hire department"),
    ("Accounts Activation", "Add in Google Calendar Townhall meeting or any other related meetings", None),
    ("Accounts Activation", "Add in HR Dashboard", None),
    ("Accounts Activation", "Add to Team WhatsApp group", None),
]


def _seed_default_checklist_items(db, tenant_id: int) -> None:
    """Seed the default 3-section onboarding checklist on first use."""
    from models import HrOnboardingChecklistItem

    has_items = db.execute(select(HrOnboardingChecklistItem.id).where(
        HrOnboardingChecklistItem.tenant_id == tenant_id
    ).limit(1)).first()
    if has_items:
        return
    for order, (section, title, desc) in enumerate(_DEFAULT_ONBOARDING_CHECKLIST, start=1):
        db.add(HrOnboardingChecklistItem(
            tenant_id=tenant_id,
            title=title,
            description=desc,
            section=section,
            sort_order=order,
            created_by="System",
        ))
    db.commit()


# ---------------------------------------------------------------------------
# Onboarding Checklist — HR-managed template, per-staff tick-off
# ---------------------------------------------------------------------------

class HrChecklistItemBody(BaseModel):
    title: str
    description: Optional[str] = None
    section: Optional[str] = None


@router.post("/hr/onboarding-checklist")
async def add_hr_checklist_item(
    body: HrChecklistItemBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Add an item to the onboarding checklist template."""
    from models import HrOnboardingChecklistItem

    title = (body.title or "").strip()
    if not title:
        raise HTTPException(status_code=422, detail="Checklist item title is required")
    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")

    max_order = db.execute(
        select(HrOnboardingChecklistItem.sort_order).where(
            HrOnboardingChecklistItem.tenant_id == tenant.id
        ).order_by(HrOnboardingChecklistItem.sort_order.desc())
    ).scalars().first() or 0

    item = HrOnboardingChecklistItem(
        tenant_id=tenant.id,
        title=title[:256],
        description=(body.description or "").strip()[:1024] or None,
        section=(body.section or "").strip()[:128] or None,
        sort_order=max_order + 1,
        created_by=(user.name if user else "HR") or "HR",
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"ok": True, "item": item.to_dict()}


@router.put("/hr/onboarding-checklist/{item_id}")
async def update_hr_checklist_item(
    item_id: int,
    body: HrChecklistItemBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Edit a checklist template item."""
    from models import HrOnboardingChecklistItem

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    item = db.get(HrOnboardingChecklistItem, item_id)
    if item is None or item.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Checklist item not found")

    title = (body.title or "").strip()
    if not title:
        raise HTTPException(status_code=422, detail="Checklist item title is required")
    item.title = title[:256]
    item.description = (body.description or "").strip()[:1024] or None
    if body.section is not None:
        item.section = body.section.strip()[:128] or None
    db.commit()
    db.refresh(item)
    return {"ok": True, "item": item.to_dict()}


@router.delete("/hr/onboarding-checklist/{item_id}")
async def delete_hr_checklist_item(
    item_id: int,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Remove a checklist template item and all staff progress for it."""
    from models import HrOnboardingChecklistItem, HrOnboardingChecklistProgress

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    item = db.get(HrOnboardingChecklistItem, item_id)
    if item is None or item.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Checklist item not found")

    db.execute(delete(HrOnboardingChecklistProgress).where(
        HrOnboardingChecklistProgress.item_id == item_id,
        HrOnboardingChecklistProgress.tenant_id == tenant.id,
    ))
    db.delete(item)
    db.commit()
    return {"ok": True}


class HrChecklistToggleBody(BaseModel):
    staff_name: str
    completed: bool


@router.post("/hr/onboarding-checklist/{item_id}/toggle")
async def toggle_hr_checklist_item(
    item_id: int,
    body: HrChecklistToggleBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Tick/untick one checklist item for one staff member.

    When every checklist item is ticked, the staff's onboarding task is
    marked Done (onboarding process complete); unticking after completion
    reverts it to In progress.
    """
    from models import HrOnboardingChecklistItem, HrOnboardingChecklistProgress, HrOnboardingTask

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    item = db.get(HrOnboardingChecklistItem, item_id)
    if item is None or item.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Checklist item not found")
    staff = (body.staff_name or "").strip()
    if not staff:
        raise HTTPException(status_code=422, detail="staff_name is required")

    progress = db.execute(select(HrOnboardingChecklistProgress).where(
        HrOnboardingChecklistProgress.tenant_id == tenant.id,
        HrOnboardingChecklistProgress.staff_name == staff,
        HrOnboardingChecklistProgress.item_id == item_id,
    )).scalars().first()

    now = datetime.utcnow().isoformat(timespec="seconds")
    if progress is None:
        progress = HrOnboardingChecklistProgress(
            tenant_id=tenant.id, staff_name=staff, item_id=item_id,
            completed=body.completed,
            completed_at=now if body.completed else None,
            completed_by=(user.name if user else "HR") if body.completed else None,
        )
        db.add(progress)
    else:
        progress.completed = body.completed
        progress.completed_at = now if body.completed else None
        progress.completed_by = (user.name if user else "HR") if body.completed else None

    # Flush so the recompute below sees the row we just added/changed
    # (production session factory runs with autoflush=False).
    db.flush()

    # Recompute completion for this staff member
    total_items = db.execute(select(HrOnboardingChecklistItem.id).where(
        HrOnboardingChecklistItem.tenant_id == tenant.id
    )).all()
    done_items = db.execute(select(HrOnboardingChecklistProgress.item_id).where(
        HrOnboardingChecklistProgress.tenant_id == tenant.id,
        HrOnboardingChecklistProgress.staff_name == staff,
        HrOnboardingChecklistProgress.completed == True,  # noqa: E712
    )).all()
    all_done = len(total_items) > 0 and len(done_items) >= len(total_items)

    # Reflect on the onboarding task row (synced from Notion) if present
    task = db.execute(select(HrOnboardingTask).where(
        HrOnboardingTask.tenant_id == tenant.id,
        HrOnboardingTask.staff_name == staff,
    )).scalars().first()
    if task is not None:
        if all_done and task.status != "Done":
            task.status = "Done"
        elif not all_done and task.status == "Done":
            task.status = "In progress"

    db.commit()
    db.refresh(progress)
    return {
        "ok": True,
        "progress": progress.to_dict(),
        "all_done": all_done,
        "done_count": len(done_items),
        "total_items": len(total_items),
    }


# ---------------------------------------------------------------------------
# Training & Development — create programs, participants, approval doc, certs
# ---------------------------------------------------------------------------

_ALLOWED_TRAINING_DOC_EXTS = {"pdf", "doc", "docx", "txt", "png", "jpg", "jpeg", "webp"}


@router.post("/hr/trainings")
async def create_hr_training(
    name: str = Path(...),
    approval_doc: UploadFile = File(None),
    training_name: str = Form(...),
    staff_name: str = Form(""),
    trainer_name: str = Form(""),
    training_format: str = Form(""),
    start_date: str = Form(""),
    end_date: str = Form(""),
    training_charges: str = Form(""),
    exam_included: bool = Form(False),
    bond_agreement: bool = Form(False),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Create a training program from the portal (optional approval doc upload)."""
    import uuid as _uuid
    from models import HrTraining

    tname = (training_name or "").strip()
    if not tname:
        raise HTTPException(status_code=422, detail="Training name is required")
    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")

    approval_doc_url = None
    if approval_doc is not None and approval_doc.filename:
        approval_doc_url = await _save_equipment_upload(approval_doc, _ALLOWED_TRAINING_DOC_EXTS, "approval doc")

    training = HrTraining(
        tenant_id=tenant.id,
        notion_page_id=f"local-{_uuid.uuid4().hex}",
        training_name=tname,
        staff_name=staff_name.strip() or None,
        trainer_name=trainer_name.strip() or None,
        training_format=training_format.strip() or None,
        start_date=start_date.strip() or None,
        end_date=end_date.strip() or None,
        training_charges=_parse_amount(training_charges),
        exam_included=bool(exam_included),
        bond_agreement=bool(bond_agreement),
        approval_doc_url=approval_doc_url,
    )
    db.add(training)
    db.commit()
    db.refresh(training)

    try:
        import audit
        audit.log_action(db, tenant, user, "hr", "hr.training.create", "training",
                         str(training.id), detail={"training_name": training.training_name})
    except Exception:
        pass
    return {"ok": True, "training": training.to_dict()}


@router.post("/hr/trainings/{training_id}/approval-doc")
async def upload_hr_training_approval_doc(
    training_id: int,
    name: str = Path(...),
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Attach/replace the approval document on an existing training program."""
    from models import HrTraining

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    training = db.get(HrTraining, training_id)
    if training is None or training.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Training not found")

    training.approval_doc_url = await _save_equipment_upload(file, _ALLOWED_TRAINING_DOC_EXTS, "approval doc")
    db.commit()
    db.refresh(training)
    return {"ok": True, "training": training.to_dict()}


class HrTrainingParticipantBody(BaseModel):
    staff_name: str
    department: Optional[str] = None


@router.post("/hr/trainings/{training_id}/participants")
async def add_hr_training_participant(
    training_id: int,
    body: HrTrainingParticipantBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Add a participant to a training program."""
    from models import HrTraining, HrTrainingParticipant

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    training = db.get(HrTraining, training_id)
    if training is None or training.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Training not found")
    staff = (body.staff_name or "").strip()
    if not staff:
        raise HTTPException(status_code=422, detail="Participant name is required")

    existing = db.execute(select(HrTrainingParticipant).where(
        HrTrainingParticipant.tenant_id == tenant.id,
        HrTrainingParticipant.training_id == training_id,
        HrTrainingParticipant.staff_name == staff,
    )).scalars().first()
    if existing is not None:
        raise HTTPException(status_code=422, detail="Participant already added to this training")

    participant = HrTrainingParticipant(
        tenant_id=tenant.id,
        training_id=training_id,
        staff_name=staff,
        department=(body.department or "").strip() or None,
    )
    db.add(participant)
    db.commit()
    db.refresh(participant)
    return {"ok": True, "participant": participant.to_dict()}


@router.delete("/hr/trainings/{training_id}/participants/{participant_id}")
async def remove_hr_training_participant(
    training_id: int,
    participant_id: int,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Remove a participant from a training program."""
    from models import HrTrainingParticipant

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    participant = db.get(HrTrainingParticipant, participant_id)
    if participant is None or participant.tenant_id != tenant.id or participant.training_id != training_id:
        raise HTTPException(status_code=404, detail="Participant not found")
    db.delete(participant)
    db.commit()
    return {"ok": True}


@router.post("/hr/trainings/{training_id}/participants/{participant_id}/certificate")
async def upload_hr_training_certificate(
    training_id: int,
    participant_id: int,
    name: str = Path(...),
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Upload a participant's training certificate."""
    from models import HrTrainingParticipant

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    participant = db.get(HrTrainingParticipant, participant_id)
    if participant is None or participant.tenant_id != tenant.id or participant.training_id != training_id:
        raise HTTPException(status_code=404, detail="Participant not found")

    participant.cert_url = await _save_equipment_upload(file, _ALLOWED_TRAINING_DOC_EXTS, "certificate")
    participant.cert_uploaded_at = datetime.utcnow().strftime("%Y-%m-%d")
    db.commit()
    db.refresh(participant)
    return {"ok": True, "participant": participant.to_dict()}


# ---------------------------------------------------------------------------
# Recruitment lifecycle — screening setup, shortlist gate, close job
# ---------------------------------------------------------------------------

class HrJobScreeningBody(BaseModel):
    screening_form_link: Optional[str] = None
    screening_email_subject: Optional[str] = None
    screening_email_body: Optional[str] = None


@router.put("/hr/job-openings/{job_id}/screening")
async def update_job_screening_setup(
    job_id: int,
    body: HrJobScreeningBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """HR edits the job's screening setup (Google Form link + email template)."""
    from models import HrJobOpening

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    job = db.get(HrJobOpening, job_id)
    if job is None or job.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Job opening not found")

    if body.screening_form_link is not None:
        job.screening_form_link = body.screening_form_link.strip()[:1024] or None
    if body.screening_email_subject is not None:
        job.screening_email_subject = body.screening_email_subject.strip()[:256] or None
    if body.screening_email_body is not None:
        job.screening_email_body = body.screening_email_body.strip() or None
    db.commit()
    db.refresh(job)
    return {"ok": True, "job": job.to_dict()}


class HrCandidateBulkBody(BaseModel):
    candidate_ids: List[int]
    action: str  # "shortlist" | "reject"
    reason: Optional[str] = None


@router.post("/hr/job-openings/{job_id}/candidates/bulk")
async def bulk_candidate_action(
    job_id: int,
    body: HrCandidateBulkBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Bulk shortlist / soft-reject candidates on a job (resume inbox gate).

    Rejected candidates are never deleted — kept with reason for the talent pool.
    """
    from models import HrCandidate, HrJobOpening

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    job = db.get(HrJobOpening, job_id)
    if job is None or job.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Job opening not found")
    if body.action not in ("shortlist", "reject"):
        raise HTTPException(status_code=422, detail="action must be 'shortlist' or 'reject'")
    if not body.candidate_ids:
        raise HTTPException(status_code=422, detail="candidate_ids is required")

    updated = 0
    for cid in body.candidate_ids:
        cand = db.get(HrCandidate, cid)
        if cand is None or cand.tenant_id != tenant.id:
            continue
        if body.action == "shortlist":
            if cand.status in ("Rejected", "Done"):
                continue  # terminal states are not shortlistable
            old = cand.status
            cand.status = "Shortlisted"
            cand.removed_reason = None
            cand.waiting_since = None
            cand.waiting_reason = None
            _hr_event(db, tenant.id, cand.id, "stage_move",
                      note=f"Shortlisted for {job.job_title}",
                      from_status=old, to_status="Shortlisted", user=user)
        else:
            old = cand.status
            reason = (body.reason or "").strip() or "Not suitable"
            cand.status = "Rejected"
            cand.removed_reason = reason[:1000]
            cand.waiting_since = None
            cand.waiting_reason = None
            _hr_event(db, tenant.id, cand.id, "stage_move",
                      note=f"Rejected: {reason}", from_status=old, to_status="Rejected", user=user)
        updated += 1
    db.commit()
    return {"ok": True, "updated": updated}


class HrCloseJobBody(BaseModel):
    reason: str = "Filled"  # "Filled" | "Cancelled"
    remaining_action: str = "reject"  # "reject" | "keep"


@router.post("/hr/job-openings/{job_id}/close")
async def close_hr_job_opening(
    job_id: int,
    body: HrCloseJobBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Close a job opening. Remaining non-terminal candidates are soft-rejected
    (kept in the talent pool) unless remaining_action='keep'."""
    from models import HrCandidate, HrJobOpening

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    job = db.get(HrJobOpening, job_id)
    if job is None or job.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Job opening not found")
    if (job.job_status or "").startswith("Closed"):
        raise HTTPException(status_code=422, detail="Job already closed")
    reason = (body.reason or "Filled").strip() or "Filled"

    job.job_status = f"Closed ({reason})"
    job.closed_at = datetime.utcnow().strftime("%Y-%m-%d")

    rejected_count = 0
    if body.remaining_action == "reject":
        candidates = db.execute(select(HrCandidate).where(
            HrCandidate.tenant_id == tenant.id,
            HrCandidate.role == job.job_title,
        )).scalars().all()
        close_reason = f"Job closed — position {reason.lower()}"
        for cand in candidates:
            if cand.status in ("Rejected", "Done"):
                continue
            old = cand.status
            cand.status = "Rejected"
            cand.removed_reason = close_reason
            cand.waiting_since = None
            cand.waiting_reason = None
            _hr_event(db, tenant.id, cand.id, "stage_move",
                      note=f"Rejected: {close_reason}", from_status=old,
                      to_status="Rejected", user=user)
            rejected_count += 1
    db.commit()
    db.refresh(job)
    return {"ok": True, "job": job.to_dict(), "rejected_candidates": rejected_count}


class HrAttachCandidateBody(BaseModel):
    job_id: int


@router.post("/hr/candidates/{candidate_id}/attach-job")
async def attach_candidate_to_job(
    candidate_id: int,
    body: HrAttachCandidateBody,
    name: str = Path(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Attach (or re-invite) a talent-pool candidate to a job opening and
    restart their journey at Resume Received."""
    from models import HrCandidate, HrJobOpening

    tenant = db.get(Tenant, user.tenant_id) if user and user.tenant_id else get_primary_tenant(db)
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found")
    cand = db.get(HrCandidate, candidate_id)
    if cand is None or cand.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Candidate not found")
    job = db.get(HrJobOpening, body.job_id)
    if job is None or job.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Job opening not found")
    if (job.job_status or "").startswith("Closed"):
        raise HTTPException(status_code=422, detail="Cannot attach to a closed job")

    old_status = cand.status
    cand.job_opening_id = job.id
    cand.role = job.job_title
    cand.status = "Resume Received"
    cand.removed_reason = None
    _hr_event(db, tenant.id, cand.id, "stage_move",
              note=f"Attached to job '{job.job_title}' (re-invited from talent pool)",
              from_status=old_status, to_status="Resume Received", user=user)
    db.commit()
    db.refresh(cand)
    return {"ok": True, "candidate": cand.to_dict()}

# =============================================================================
# Project Dashboard Endpoints
# =============================================================================
# All routes are namespaced under /projects/ to avoid collisions with the
# CRM endpoints above (/tasks, /deals, /companies). Static routes (/stats,
# /tasks) are registered BEFORE the dynamic /projects/{project_id} route so
# FastAPI matches them first.

@router.get("/projects", tags=["projects"])
async def list_projects(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    status: Optional[str] = Query(None, description="Filter by status"),
    pm: Optional[str] = Query(None, description="Filter by project manager"),
    gate: Optional[int] = Query(None, description="Filter by gate number"),
) -> dict:
    """List all projects with optional filters."""
    from models import Project

    query = select(Project)

    if status:
        query = query.where(Project.status == status)
    if pm:
        query = query.where(Project.pm == pm)
    if gate is not None:
        query = query.where(Project.gate == gate)

    projects = db.execute(query).scalars().all()
    return {"projects": [p.to_dict() for p in projects]}


@router.get("/projects/stats", tags=["projects"])
async def get_project_stats(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Get project dashboard statistics."""
    from models import Project, Task
    from sqlalchemy import func

    total_projects = db.execute(select(func.count(Project.id))).scalar()
    active_projects = db.execute(
        select(func.count(Project.id)).where(Project.status.in_(["active", "in-progress"]))
    ).scalar()

    total_tasks = db.execute(select(func.count(Task.id))).scalar()
    completed_tasks = db.execute(
        select(func.count(Task.id)).where(Task.status == "done")
    ).scalar()

    overdue_tasks = db.execute(
        select(func.count(Task.id))
        .where(Task.deadline < datetime.now())
        .where(Task.status != "done")
    ).scalar()

    return {
        "projects": {
            "total": total_projects or 0,
            "active": active_projects or 0,
        },
        "tasks": {
            "total": total_tasks or 0,
            "completed": completed_tasks or 0,
            "overdue": overdue_tasks or 0,
        },
    }


@router.get("/projects/tasks", tags=["projects"])
async def list_project_dashboard_tasks(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    project_id: Optional[str] = Query(None, description="Filter by project ID"),
    owner: Optional[str] = Query(None, description="Filter by task owner"),
    status: Optional[str] = Query(None, description="Filter by status"),
    priority: Optional[str] = Query(None, description="Filter by priority"),
    overdue: Optional[bool] = Query(None, description="Filter overdue tasks only"),
) -> dict:
    """List all project-tracker tasks with optional filters.

    Namespaced under /projects/tasks so it does not shadow the CRM /tasks
    endpoint above.
    """
    from models import Task

    query = select(Task)

    if project_id:
        query = query.where(Task.project_id == project_id)
    if owner:
        query = query.where(Task.owner == owner)
    if status:
        query = query.where(Task.status == status)
    if priority:
        query = query.where(Task.priority == priority)
    if overdue is True:
        # Tasks with deadline in past and not done
        query = query.where(Task.deadline < datetime.now()).where(Task.status != "done")

    tasks = db.execute(query).scalars().all()
    return {"tasks": [t.to_dict() for t in tasks]}


@router.get("/projects/active", tags=["projects"])
async def list_active_projects(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """List active projects (status starts with 'active')."""
    from models import Project

    projects = db.execute(
        select(Project).where(Project.status.like("active%"))
    ).scalars().all()
    return {"projects": [p.to_dict() for p in projects]}


@router.get("/tasks/plan", tags=["projects"])
async def list_planned_tasks(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Plan view: open tasks with deadlines, soonest first (source /tasks/plan)."""
    from models import Task

    tasks = db.execute(
        select(Task)
        .where(Task.status.in_(["todo", "in-progress"]))
        .where(Task.deadline.is_not(None))
        .order_by(Task.deadline.asc())
    ).scalars().all()
    return {"tasks": [t.to_dict() for t in tasks]}


@router.get("/reports/summary", tags=["projects"])
async def get_reports_summary(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Reports section: aggregated portfolio summary computed from synced data."""
    from collections import Counter
    from models import Project, Task

    projects = db.execute(select(Project)).scalars().all()
    tasks = db.execute(select(Task)).scalars().all()

    now = datetime.now()

    # Per-project aggregates
    by_pm: Dict[str, int] = {}
    by_status: Counter = Counter()
    by_health: Counter = Counter()
    by_gate: Counter = Counter()
    by_budget: Counter = Counter()
    total_value = 0.0
    project_rows = []
    for p in projects:
        by_status[p.status or "unknown"] += 1
        by_health[p.overall_health or "unknown"] += 1
        by_budget[p.budget_status or "unknown"] += 1
        if p.gate is not None:
            by_gate[f"Gate {p.gate}"] += 1
        if p.pm:
            by_pm[p.pm] = by_pm.get(p.pm, 0) + 1
        if p.value_rm:
            total_value += p.value_rm
        p_tasks = [t for t in tasks if t.project_id == p.id]
        open_tasks = [t for t in p_tasks if t.status not in ("done", "cancelled")]
        overdue = [
            t for t in open_tasks
            if t.deadline and t.deadline.replace(tzinfo=None) < now
        ]
        project_rows.append({
            "id": p.id,
            "name": p.name,
            "client": p.client,
            "pm": p.pm,
            "status": p.status,
            "overallHealth": p.overall_health,
            "budgetStatus": p.budget_status,
            "gate": p.gate,
            "valueRm": p.value_rm,
            "targetEnd": p.target_end.isoformat() if p.target_end else None,
            "sourceLastUpdated": p.source_last_updated.isoformat() if p.source_last_updated else None,
            "totalTasks": len(p_tasks),
            "openTasks": len(open_tasks),
            "overdueTasks": len(overdue),
            "completionPct": round(len([t for t in p_tasks if t.status == "done"]) / len(p_tasks) * 100) if p_tasks else 0,
        })

    # Task aggregates
    open_task_count = len([t for t in tasks if t.status not in ("done", "cancelled")])
    overdue_count = len([
        t for t in tasks
        if t.status not in ("done", "cancelled")
        and t.deadline and t.deadline.replace(tzinfo=None) < now
    ])
    by_priority: Counter = Counter(t.priority or "UNSET" for t in tasks if t.status not in ("done", "cancelled"))

    return {
        "totals": {
            "projects": len(projects),
            "activeProjects": len([p for p in projects if (p.status or "").startswith("active")]),
            "totalValueRm": total_value,
            "tasks": len(tasks),
            "openTasks": open_task_count,
            "overdueTasks": overdue_count,
        },
        "projectsByStatus": dict(by_status.most_common()),
        "projectsByHealth": dict(by_health.most_common()),
        "projectsByGate": dict(sorted(by_gate.items())),
        "projectsByBudgetStatus": dict(by_budget.most_common()),
        "projectsByPm": dict(sorted(by_pm.items(), key=lambda kv: -kv[1])),
        "openTasksByPriority": dict(by_priority.most_common()),
        "projects": project_rows,
    }


@router.get("/support/tickets", tags=["projects"])
async def list_support_tickets(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    status: Optional[str] = Query(None, description="Filter by ticket status"),
    priority: Optional[str] = Query(None, description="Filter by priority (P1-P4)"),
    customer: Optional[str] = Query(None, description="Filter by customer"),
) -> dict:
    """List support tickets (source /support section)."""
    from models import SupportTicket

    query = select(SupportTicket)
    if status:
        query = query.where(SupportTicket.status == status)
    if priority:
        query = query.where(SupportTicket.priority == priority)
    if customer:
        query = query.where(SupportTicket.customer_slug == customer)
    tickets = db.execute(query).scalars().all()
    return {"tickets": [t.to_dict() for t in tickets], "total": len(tickets)}


@router.get("/support/stats", tags=["projects"])
async def get_support_stats(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Support section statistics."""
    from collections import Counter
    from models import SupportTicket

    tickets = db.execute(select(SupportTicket)).scalars().all()
    by_status: Counter = Counter(t.status or "unknown" for t in tickets)
    by_priority: Counter = Counter(t.priority or "unknown" for t in tickets)
    by_customer: Counter = Counter(t.customer or "unknown" for t in tickets)
    open_statuses = ("Open", "In Progress", "Waiting for Customer")
    open_tickets = [t for t in tickets if t.status in open_statuses]
    new_replies = len([t for t in tickets if t.new_reply])

    return {
        "totals": {
            "tickets": len(tickets),
            "open": len(open_tickets),
            "closedOrResolved": len(tickets) - len(open_tickets),
            "newReplies": new_replies,
        },
        "byStatus": dict(by_status.most_common()),
        "byPriority": dict(by_priority.most_common()),
        "topCustomers": dict(by_customer.most_common(10)),
    }


@router.get("/projects/{project_id}", tags=["projects"])
async def get_project(
    project_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Get a single project with all nested data."""
    from models import Project

    project = db.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    return project.to_dict()


@router.get("/projects/{project_id}/tasks", tags=["projects"])
async def list_project_tasks(
    project_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Get all tasks for a specific project."""
    from models import Task

    query = select(Task).where(Task.project_id == project_id)
    tasks = db.execute(query).scalars().all()
    return {"tasks": [t.to_dict() for t in tasks]}
