#!/usr/bin/env python3
"""balance-sheet-to-snapshot.py

Read the BS sheet of ~/brain/finance/202606-management-report.xlsx and write
finance/snapshots/balance-sheet in gbrain — full asset picture including
long-term (non-current) assets, liabilities, and equity.

Why: the June infographic only described current assets; the real balance
sheet also holds RM 151,588 of long-term assets (PP&E at NBV + subsidiary
investment) and RM 2.40M total assets.

Sheet layout (indent hierarchy):
    "   Current Assets"   (section header, 3 spaces)
    "      Trade and other receivables"   (category, 6 spaces) + amount
    "         <account>"                  (sub-item, 9 spaces) + amount
    "      Total ..."                     (rollup, 6 spaces)

Writes the same AssetCategory contract the AssetTab UI renders:
  current_assets[]     {name, amount, icon, sub_items:[{name,amount}]}
  non_current_assets[] {name, amount, icon, sub_items:[...]}
  total_current_assets / total_non_current_assets / total_assets
  total_liabilities / total_current_liabilities / total_equity
  asset_trend[]  (single month point — no fabricated history)

Idempotent; safe to re-run after each month's management report.
"""
import json, os, re, subprocess
from datetime import datetime
import openpyxl

XLSX = os.environ.get("FINANCE_REPORT_XLSX", "/home/tapway/brain/finance/202606-management-report.xlsx")
PGPW = os.environ.get("GBRAIN_PG_PASSWORD")
if not PGPW:
    raise RuntimeError("GBRAIN_PG_PASSWORD environment variable is required but not set")
SLUG = "finance/snapshots/balance-sheet"
PERIOD = "2026-06"

ICONS = {
    "trade": "FileText", "cash": "Wallet", "inventory": "Package",
    "prepaid": "CalendarClock", "deposit": "Landmark", "retention": "ShieldCheck",
    "expense": "TrendingDown", "computer": "Brain", "furniture": "Building2",
    "motor": "Layers", "office": "Layers", "renovation": "Building2",
    "investment": "Landmark",
}


def name_icon(name: str) -> str:
    n = name.lower()
    for key, icon in ICONS.items():
        if key in n:
            return icon
    return "Wallet"


# NOTE: out_rollups param is unused — consider removing for clarity
def parse_bs(path, out_rollups):
    """Parse the BS sheet of one management-report xlsx into rollup buckets."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "BS" not in wb.sheetnames:
        return None
    ws = wb["BS"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    sections = {}          # section label → list of (category, amount, [subs])
    section = None
    current_cat = None
    rollups = {}

    for r in rows:
        name = str(r[0] or "").strip()
        if not name:
            continue
        low = name.lower()
        amt = float(r[1]) if isinstance(r[1], (int, float)) else 0.0

        # ── rollup / total rows FIRST (they also contain section words) ──
        if low.startswith("total "):
            rollups[low] = amt
            if current_cat is not None:
                tname = "total " + current_cat["name"].lower()
                if low == tname:
                    current_cat["amount"] = amt  # net value (after dep)
                    current_cat = None
            continue

        # ── section headers ──
        if name.strip().rstrip(":") in ("Assets",) or (
            "liabilities and" in low and "equity" in low
        ):
            section = "assets"; sections.setdefault(section, []); current_cat = None; continue
        if low in ("current assets",):
            section = "current_assets"; sections.setdefault(section, []); current_cat = None; continue
        if low in ("long-term assets", "long term assets", "long-term assets:", "long term assets:"):
            section = "long_term_assets"; sections.setdefault(section, []); current_cat = None; continue
        if low in ("current liabilities:", "current liabilities"):
            section = "current_liabilities"; sections.setdefault(section, []); current_cat = None; continue
        if low in ("non-current liabilities:", "non-current liabilities", "non current liabilities:", "non current liabilities"):
            section = "non_current_liabilities"; sections.setdefault(section, []); current_cat = None; continue
        if "shareholder" in low:
            section = "equity"; sections.setdefault(section, []); current_cat = None; continue
        if section is None:
            continue

        # Normalize tabs to spaces for consistent indent detection
        row_label = str(r[0] or "").expandtabs(4)
        indent = len(row_label) - len(row_label.lstrip(" "))

        # category vs sub-item by indent
        if indent >= 9 and current_cat is not None:
            current_cat["sub_items"].append({"name": name, "amount": amt})
        else:
            cat = {"name": name, "amount": amt, "icon": name_icon(name), "sub_items": []}
            sections[section].append(cat)
            current_cat = cat

    def find_rollup(*keys):
        for k in keys:
            if k in rollups:
                return rollups[k]
        return 0.0

    return {
        "current_assets": sections.get("current_assets", []),
        "non_current_assets": sections.get("long_term_assets", []),
        "total_current_assets": find_rollup("total current assets"),
        "total_non_current_assets": find_rollup("total long-term assets", "total long term assets"),
        "total_assets": find_rollup("total assets"),
        "total_current_liabilities": find_rollup("total current liabilities"),
        "total_liabilities": find_rollup("total non-current liabilities", "total non current liabilities"),
        "total_equity": find_rollup("total shareholders' equity", "total shareholders equity"),
    }


def find_reports():
    """All monthly management reports under the finance dir."""
    found = []
    for root, _dirs, files in os.walk(os.path.dirname(XLSX)):
        for f in sorted(files):
            if "management-report" in f.lower() and f.lower().endswith(".xlsx"):
                found.append(os.path.join(root, f))
    return found


def month_tag(path: str) -> str:
    """'2026-06' from a report filename like 202606-management-report.xlsx."""
    base = os.path.basename(path)
    m = re.search(r"(\d{4})(\d{2})", base)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    # fall back to the xlsx mtime
    return datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m")


def parse(wb):
    # gather the LATEST report's full detail (this report)
    data = parse_bs(XLSX, None)
    if data is None:
        raise SystemExit(f"no BS sheet in {XLSX}")

    def clean_cats(items):
        out = []
        for c in items:
            c = dict(c)
            c["sub_items"] = [s for s in c.get("sub_items", []) if abs(s["amount"]) > 0.005]
            out.append(c)
        return out

    # ── build the asset trend from ALL reports (oldest → newest) ──
    trend = []
    for path in find_reports():
        d = parse_bs(path, None)
        if d is None:
            continue
        if not d["total_assets"]:
            d["total_assets"] = round(d["total_current_assets"] + d["total_non_current_assets"], 2)
        trend.append({
            "month": month_tag(path),
            "current": round(d["total_current_assets"], 2),
            "non_current": round(d["total_non_current_assets"], 2),
        })
    trend.sort(key=lambda x: x["month"])

    total_lt = data["total_non_current_assets"]
    if not data["total_assets"]:
        data["total_assets"] = round(data["total_current_assets"] + total_lt, 2)
    total_liab = round(data["total_current_liabilities"] + data["total_liabilities"], 2)

    return {
        "current_assets": clean_cats(data["current_assets"]),
        "non_current_assets": clean_cats(data["non_current_assets"]),
        "total_current_assets": round(data["total_current_assets"], 2),
        "total_non_current_assets": round(total_lt, 2),
        "total_assets": round(data["total_assets"], 2),
        "total_current_liabilities": round(data["total_current_liabilities"], 2),
        "total_liabilities": round(total_liab, 2),
        "total_equity": round(data["total_equity"], 2),
        "asset_trend": trend,
        "period": PERIOD,
        "source": os.path.basename(XLSX),
    }


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    data = parse(wb)

    fm = json.dumps(data, default=str).replace("'", "''")
    body = json.dumps(data, indent=2, default=str).replace("'", "''")
    sql = f"""
    INSERT INTO pages (source_id, slug, type, page_kind, title, frontmatter, compiled_truth,
                       content_hash, created_at, updated_at)
    VALUES ('default', '{SLUG}', 'finance', 'markdown', 'Finance Snapshot — Balance Sheet (Jun 2026)',
            '{fm}'::jsonb, '{body}', '', now(), now())
    ON CONFLICT (source_id, slug) DO UPDATE SET
        frontmatter = EXCLUDED.frontmatter,
        compiled_truth = EXCLUDED.compiled_truth,
        updated_at = now();
    """
    env = dict(os.environ); env["PGPASSWORD"] = PGPW
    p = subprocess.run(["psql", "-h", "127.0.0.1", "-U", "hermes", "-d", "gbrain", "-c", sql],
                       capture_output=True, text=True, env=env, timeout=60)
    if p.returncode != 0:
        raise SystemExit(p.stderr[-400:])

    print(f"✓ {SLUG} updated from {os.path.basename(XLSX)} (BS sheet)")
    print(f"  Total Assets      RM {data['total_assets']:>14,.2f}")
    print(f"  Current Assets    RM {data['total_current_assets']:>14,.2f}  ({len(data['current_assets'])} categories)")
    print(f"  Non-Current Assets RM {data['total_non_current_assets']:>13,.2f}  ({len(data['non_current_assets'])} categories)")
    print(f"  Current Liab.     RM {data['total_current_liabilities']:>14,.2f}")
    print(f"  Total Liabilities RM {data['total_liabilities']:>14,.2f}")
    print(f"  Total Equity      RM {data['total_equity']:>14,.2f}")
    print("  Non-current breakdown:")
    for c in data["non_current_assets"]:
        print(f"    • {c['name']:<32} RM {c['amount']:>12,.2f}"
              + (f"  ({len(c['sub_items'])} accounts)" if c["sub_items"] else ""))


if __name__ == "__main__":
    main()